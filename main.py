# -*- coding: utf-8 -*-

"""
Наряд-Заказ — v3.0
Исправления и улучшения:
- Онлайн-синхронизация компаний через Google Sheets
- Автоматическое обновление списка компаний при изменениях
- Добавлена поддержка НДС для компаний
- Цены для компаний с НДС берутся из отдельных файлов price_nds.xlsx и consumables_nds.xlsx
- В админ-панели добавлен переключатель НДС для компаний
- При создании новой компании НДС по умолчанию "нет"
- Добавлен выбор типа колес для грузовых и легковых автомобилей
- Цены теперь берутся с учетом выбранного типа колес
- Админ-операции (добавить компанию/номер, выставить оплату, удалить) работают даже когда окно «Создать наряд» НЕ открыто.
- Больше нет ошибок invalid command name при обновлении списков.
- Тумблер «Оплата» в админке корректно отражает состояние из файла и сразу обновляется при выборе компании/поиске.
- Компании в списках идут в том же порядке, что и в файле; новые добавляются В КОНЕЦ.
- Добавлена возможность выбора и добавления колес в наряд-заказ
Исправления от пользователя:
1. Убрано поле "Фамилия механика" (подпись ставится вручную в документе)
2. Дефект можно оставить пустым (добавлена опция "Пропустить")
3. Цены динамически обновляются при выборе типа колеса
4. Разделение на две страницы: Page 1 (клиентские данные) и Page 2 (услуги и формирование)
"""

import os
import datetime
import subprocess
from pathlib import Path
import tkinter as tk
from tkinter import BOTH, LEFT, RIGHT, Y, X, NW, DISABLED, NORMAL, messagebox, simpledialog
from tkinter import ttk

from openpyxl import load_workbook
from num2words import num2words
import pandas as pd
import ttkbootstrap as tb
from openpyxl.styles import Alignment
# Добавляем импорт для работы с временем
import time
from datetime import datetime as dt  # Импортируем с алиасом
import datetime  # Оставляем для timedelta
import glob

# === Пути проекта ===
BASE_DIR = Path(__file__).parent
TEMPLATES_DIR = BASE_DIR / "templates"
OUTPUT_BASE_DIR = BASE_DIR / "output"  # Основная папка output
DATA_DIR = BASE_DIR / "data"
TEMPLATE_XLSX = TEMPLATES_DIR / "order_template.xlsx"
COMPANIES_XLSX = DATA_DIR / "companies.xlsx"
PRICE_XLSX = DATA_DIR / "price.xlsx"
PRICE_NDS_XLSX = DATA_DIR / "price_nds.xlsx"
CONSUMABLES_XLSX = DATA_DIR / "consumables.xlsx"
CONSUMABLES_NDS_XLSX = DATA_DIR / "consumables_nds.xlsx"
WHEELS_XLSX = DATA_DIR / "wheels.xlsx"  # НОВЫЙ ФАЙЛ ДЛЯ КОЛЕС

OUTPUT_BASE_DIR.mkdir(exist_ok=True, parents=True)
TEMPLATES_DIR.mkdir(exist_ok=True, parents=True)
DATA_DIR.mkdir(exist_ok=True, parents=True)

# === Ячейки шаблона ===
CELL_CUSTOMER = "I5"
CELL_PLATE = "G6"
CELL_DRIVER = "G7"

CELL_DEFECT_LINE1 = "Y8"
CELL_DEFECT_LINE2 = "A9"
CELL_ISSUED_TO = "N10"
CELL_DATE = "CG4"
# Итоговые суммы и подпис
CELL_TOTAL_NUM = "BR41"  # ИЗМЕНЕНО: Итоговая сумма в BR41:CO41
CELL_TOTAL_TEXT = "A43"  # ИЗМЕНЕНО: Сумма прописью в A43:BE43
# Верхняя левая ячейка объединённого диапазона для механика
CELL_MECHANIC = "W52"

SERVICES_START_ROW = 13
COL_QTY = "BF"
COL_PRICE = "BR"
COL_COST = "CD"

# Диапазон для добавления колес
WHEELS_START_ROW = 38  # ИЗМЕНЕНО: Автошины начинаются с 38 строки
WHEELS_START_COL = "BF"
WHEELS_END_COL = "BQ"


def load_wheels_from_excel() -> list[str]:
    """Загружает список колес из Excel файла"""
    wheels_list = []

    if not WHEELS_XLSX.exists():
        # Если файла нет, создаем его с дефолтными значениями
        print(f"Файл с колесами не найден. Создаю {WHEELS_XLSX}")
        default_wheels = [
            "215/75R17.5 КАМА NF 202",
            "215/75R17.5 КАМА NR 201",
            "215/75R17.5 КАМА NU 301",
            "235/75R17.5 КАМА NT 202",
            "245/70R19.5 КАМА NF 201",
            "245/70R19.5 КАМА NF 202",
            "245/70R19.5 КАМА NR 201",
            "245/70R19.5 KAMA NU 301",
            "295/80R22.5 KAMA NR 202",
            "295/80R22.5 KAMA NF 202",
            "315/80R22.5 KAMA NF 201",
            "315/60R22.5 КАМА NF 202",
            "315/60R22.5 КАМА NR 201",
            "315/70R22.5 КАМА NF 202",
            "315/70R22.5 КАМА NR 202",
            "315/70R22.5 КАМА PRO NF 203",
            "385/55R22.5 КАМА NT 202",
            "385/55R22.5 КАМА PRO NF 203",
            "385/65R22.5 FORZAREG T",
            "385/65R22.5 КАМА NT 201",
            "385/65R22.5 КАМА NT 202",
            "385/65R22.5 КАМА PRO NT 203",
            "385/65R22.5 КАМА NF 202",
            "385/65R22.5 КАМА PRO NF 203",
            "10.00R20 КАМА-310",
            "Диск 9.00-22.5 10/335 D281 ET175 (Китай)",
            "А/диск 11.75-22.5 10*335 Et0",
            "А/диск 11.75-22.5 10*335 Et120",
            "Диск колеса 11,75 R22,5 SRW 10/335 ET135 D281(руль)"
        ]

        # Создаем DataFrame и сохраняем в Excel
        df = pd.DataFrame({"Колеса": default_wheels})
        df.to_excel(WHEELS_XLSX, index=False)
        return default_wheels

    try:
        # Читаем Excel файл
        df = pd.read_excel(WHEELS_XLSX, dtype=str)

        # Проверяем наличие нужной колонки
        if "Колеса" in df.columns:
            wheels_list = df["Колеса"].dropna().astype(str).tolist()
        else:
            # Если нет колонки "Колеса", используем первую колонку
            wheels_list = df.iloc[:, 0].dropna().astype(str).tolist()

        # Очищаем от лишних пробелов
        wheels_list = [wheel.strip() for wheel in wheels_list if wheel.strip()]

        # Если список пустой, возвращаем дефолтные значения
        if not wheels_list:
            print("Файл с колесами пустой. Использую значения по умолчанию.")
            wheels_list = ["215/75R17.5 КАМА NF 202", "215/75R17.5 КАМА NR 201"]

    except Exception as e:
        print(f"Ошибка загрузки колес из Excel: {e}")
        # Возвращаем дефолтные значения в случае ошибки
        wheels_list = ["215/75R17.5 КАМА NF 202", "215/75R17.5 КАМА NR 201"]

    return wheels_list

DEFECTS = [
    "Пропустить",  # ← ИЗМЕНЕНО: добавлено первым
    "Износ автошины",
    "Повреждение автошины",
    "Деформация (грыжа)",
    "Искажение протектора",
    "Трещина на боковой части шины",
    "Вмятина на протекторе",
    "Расслоение и деформация протектора",
    "Разрыв протектора",
    "Разрыв по боковине",
    "Механический разрез боковина",
    "Установка новых автошин",
    "Сезонная перебортировка колёс",
    "Вулканизация",
    "Накачка шин",
    "Другое (ввести вручную)",
]

# ИЗМЕНЕН СПИСОК УСЛУГ В СООТВЕТСТВИИ С ШАБЛОНОМ
SERVICES = [
    "Снятие/установка",
    "Мойка колёс",
    "Разбортовка",
    "Забортовка",
    "Балансировка",
    "Установка камеры",
    "Ремонт камеры",
    "Герметик",
    "Вентиль грузовой",
    "Вентиль ремонтный",
    "Вентиль легковой",
    "Грибок №",
    "Грузики",
    "Удлинитель",
    "Установка вентиля",
    "Утилизация",
    "Подкачка",
    "Жгут",
    "Разгрузка и погрузка колеса",
    "Косметическая варка",
    "Пластырь №",
    "Нарезка протектора одна дорожка",
    "Протяжка колёс",
    "Проверка на герметичность",
    "Упаковочный пакет",
    # Установленные автошины НЕ являются услугой - они заполняются отдельно
]

WHEELS_LIST = load_wheels_from_excel()

# === Настройки Google Sheets ===
GOOGLE_SHEET_ID = "1FMjGjD1ZUI7EyixhXVmFFN2iQTJua_Lm84Q9yLSxkbA"  # Замените на ID вашей Google таблицы
GOOGLE_SHEET_NAME = "Sheet1"  # Название листа в Google таблице
GOOGLE_CREDENTIALS_FILE = BASE_DIR / "credentials.json"  # Файл с учетными данными

# === Работа с компаниями ===
COL_NAME = "Компания"
COL_INN = "ИНН"
COL_PLATES = "Номера"
COL_PAY = "Оплата"
COL_VAT = "НДС"


# Функция для получения текущей папки дня
def get_current_day_folder() -> Path:
    """
    Возвращает папку для текущего дня.
    Создает новую папку в 07:00 по МСК каждый день.
    """
    # Получаем текущее время в UTC
    utc_now = dt.utcnow()

    # МСК = UTC+3
    moscow_offset = datetime.timedelta(hours=3)
    moscow_time = utc_now + moscow_offset

    # Проверяем, нужно ли создать новую папку (после 07:00 по МСК)
    target_hour = 7  # 07:00 по МСК

    # Если сейчас до 07:00, используем вчерашнюю дату
    if moscow_time.hour < target_hour:
        folder_date = moscow_time - datetime.timedelta(days=1)
    else:
        folder_date = moscow_time

    # Форматируем дату для имени папки
    folder_name = folder_date.strftime("%Y-%m-%d")
    day_folder = OUTPUT_BASE_DIR / folder_name

    # Создаем папку если она не существует
    day_folder.mkdir(exist_ok=True, parents=True)

    return day_folder

def _normalize_company_df(df: pd.DataFrame) -> pd.DataFrame:
    # Поддержка разных заголовков (включая варианты вроде "Оплата (да/нет)")
    mapping = {}
    for col in df.columns:
        v = str(col).strip().lower()
        if v in ("компания", "название", "организация", "контрагент", "company", "name"):
            mapping[col] = COL_NAME
        elif v in ("инн", "inn"):
            mapping[col] = COL_INN
        elif v in ("номера", "госномер", "госномера", "машины", "авто", "plates", "cars"):
            mapping[col] = COL_PLATES
        elif ("оплат" in v) or v in ("оплата", "опл", "pay", "payment"):
            mapping[col] = COL_PAY
        elif v in ("ндс", "vat", "ндс"):
            mapping[col] = COL_VAT
    df2 = df.rename(columns=mapping).copy()
    for c in (COL_NAME, COL_INN, COL_PLATES, COL_PAY, COL_VAT):
        if c not in df2.columns:
            df2[c] = ""
    df2 = df2[[COL_NAME, COL_INN, COL_PLATES, COL_PAY, COL_VAT]]
    for c in (COL_NAME, COL_INN, COL_PLATES, COL_PAY, COL_VAT):
        df2[c] = df2[c].astype(str).fillna("").str.strip()
    return df2


def read_companies_df() -> pd.DataFrame:
    """Читает компании из Google Sheets или локального файла"""
    try:
        # Пробуем прочитать из Google Sheets
        df = read_companies_from_google()
        if df is not None:
            # Сохраняем локальную копию
            df.to_excel(COMPANIES_XLSX, index=False)
            return _normalize_company_df(df)
    except Exception as e:
        print(f"Ошибка чтения из Google Sheets: {e}")

    # Если Google Sheets недоступен, читаем из локального файла
    try:
        df = pd.read_excel(COMPANIES_XLSX, dtype=str)
        return _normalize_company_df(df)
    except Exception:
        # Создаем пустой DataFrame если файла нет
        return pd.DataFrame(columns=[COL_NAME, COL_INN, COL_PLATES, COL_PAY, COL_VAT])


def write_companies_df(df: pd.DataFrame):
    """Сохраняет компании в Google Sheets и локальный файл с оптимизацией"""
    try:
        # Сначала сохраняем локально для скорости
        df.to_excel(COMPANIES_XLSX, index=False)

        # Затем пробуем сохранить в Google Sheets (асинхронно, чтобы не блокировать UI)
        def sync_to_google():
            try:
                success = write_companies_to_google(df)
                if success:
                    print("✅ Синхронизация с Google Sheets завершена")
                else:
                    print("❌ Ошибка синхронизации с Google Sheets")
            except Exception as e:
                print(f"❌ Ошибка при синхронизации: {e}")

        # Запускаем в отдельном потоке, чтобы не блокировать UI
        import threading
        sync_thread = threading.Thread(target=sync_to_google, daemon=True)
        sync_thread.start()

    except Exception as e:
        print(f"Ошибка сохранения компаний: {e}")
        # Если не удалось сохранить локально, пробуем только Google Sheets
        try:
            write_companies_to_google(df)
        except Exception:
            pass


def read_companies_from_google():
    """Читает данные из Google Sheets"""
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        if not GOOGLE_CREDENTIALS_FILE.exists():
            print("Файл учетных данных Google не найден")
            return None

        # Настройка аутентификации
        scopes = ['https://www.googleapis.com/auth/spreadsheets']
        creds = Credentials.from_service_account_file(GOOGLE_CREDENTIALS_FILE, scopes=scopes)
        client = gspread.authorize(creds)

        # Открываем таблицу
        sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet(GOOGLE_SHEET_NAME)

        # Получаем все данные
        data = sheet.get_all_records()

        if not data:
            return pd.DataFrame(columns=[COL_NAME, COL_INN, COL_PLATES, COL_PAY, COL_VAT])

        return pd.DataFrame(data)

    except ImportError:
        print("Библиотеки gspread или google-auth не установлены")
        return None
    except Exception as e:
        print(f"Ошибка чтения из Google Sheets: {e}")
        return None


def write_companies_to_google(df: pd.DataFrame):
    """Записывает данные в Google Sheets с проверкой изменений"""
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        if not GOOGLE_CREDENTIALS_FILE.exists():
            print("Файл учетных данных Google не найден")
            return False

        # Настройка аутентификации
        scopes = ['https://www.googleapis.com/auth/spreadsheets']
        creds = Credentials.from_service_account_file(GOOGLE_CREDENTIALS_FILE, scopes=scopes)
        client = gspread.authorize(creds)

        # Открываем таблицу
        sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet(GOOGLE_SHEET_NAME)

        # Получаем текущие данные из Google Sheets для сравнения
        try:
            current_data = sheet.get_all_records()
            current_df = pd.DataFrame(current_data) if current_data else pd.DataFrame()
        except Exception as e:
            print(f"Ошибка чтения текущих данных: {e}")
            current_df = pd.DataFrame()

        # Если данные не изменились, не перезаписываем
        if not current_df.empty:
            current_df_normalized = _normalize_company_df(current_df)
            df_normalized = _normalize_company_df(df)

            # Сравниваем данные
            if _dataframes_equal(current_df_normalized, df_normalized):
                print("Данные не изменились, пропускаем запись в Google Sheets")
                return True

        # Очищаем лист только если данные изменились
        sheet.clear()

        # Записываем заголовки
        headers = [COL_NAME, COL_INN, COL_PLATES, COL_PAY, COL_VAT]
        sheet.append_row(headers)

        # Записываем данные порциями (batch update для больших объемов)
        batch_size = 50
        data_rows = []

        for _, row in df.iterrows():
            data_rows.append([row[col] for col in headers])

            # Отправляем порциями
            if len(data_rows) >= batch_size:
                sheet.append_rows(data_rows)
                data_rows = []

        # Отправляем оставшиеся данные
        if data_rows:
            sheet.append_rows(data_rows)

        return True

    except ImportError:
        print("Библиотеки gspread или google-auth не установлены")
        return False
    except Exception as e:
        print(f"Ошибка записи в Google Sheets: {e}")
        return False


def parse_plates(cell_value: str) -> list[str]:
    return [p.strip() for p in str(cell_value).split(",") if p.strip()]


def join_plates(plates: list[str]) -> str:
    return ", ".join(sorted(set([p.strip() for p in plates if p.strip()])))


def load_companies() -> tuple[dict, list[str]]:
    df = read_companies_df()
    companies = {}
    visible_names = []
    for _, row in df.iterrows():  # сохраняем порядок строк
        name = row[COL_NAME]
        inn = row[COL_INN]
        plates_all = parse_plates(row[COL_PLATES])
        cars = [p for p in plates_all if not p.lower().startswith("прицеп")]
        trailers = [p for p in plates_all if
                    p.lower().startswith("прицеп") or p.lower().startswith("полуприцеп") or p.lower().startswith("п/п")]
        pay = str(row[COL_PAY]).strip().lower()
        vat = str(row[COL_VAT]).strip().lower()
        if name:
            companies[name] = {
                "inn": inn,
                "plates": plates_all,
                "cars": cars,
                "trailers": trailers,
                "pay": pay,
                "vat": vat,
            }
            if pay in ("да", "yes", "true", "1"):
                visible_names.append(name)
    return companies, visible_names


COMPANIES, ALL_COMPANY_NAMES = load_companies()


def reload_companies_globals():
    """Перезагружает компании с обработкой ошибок"""
    global COMPANIES, ALL_COMPANY_NAMES
    try:
        COMPANIES, ALL_COMPANY_NAMES = load_companies()
        print(f"✅ Загружено {len(COMPANIES)} компаний, {len(ALL_COMPANY_NAMES)} с оплатой")
    except Exception as e:
        print(f"❌ Ошибка загрузки компаний: {e}")
        # Сохраняем текущее состояние если не удалось загрузить
        if 'COMPANIES' not in globals():
            COMPANIES, ALL_COMPANY_NAMES = {}, []


def _dataframes_equal(df1, df2):
    """Сравнивает два DataFrame с учетом возможных различий в типах данных"""
    try:
        # Приводим к одинаковым типам для сравнения
        df1 = df1.astype(str).fillna('')
        df2 = df2.astype(str).fillna('')
        return df1.equals(df2)
    except Exception:
        return False


def filter_companies(query: str) -> list[str]:
    q = str(query).strip().lower()
    if not q:
        return list(ALL_COMPANY_NAMES)
    result = []
    for name in ALL_COMPANY_NAMES:
        meta = COMPANIES.get(name, {})
        plates = meta.get("plates", [])
        if q in name.lower() or any(q in p.lower() for p in plates):
            result.append(name)
    return result


# === Типы колес из таблицы ===
def load_wheel_types():
    """Загружает типы колес из таблицы цен"""
    wheel_types = {
        "Грузовой": [],
        "Легковой": [],
        "Спецтехника": []
    }

    if not PRICE_XLSX.exists():
        return wheel_types

    try:
        wb = load_workbook(PRICE_XLSX, data_only=True)
        ws = wb.active

        # Ищем строку с заголовками для грузовых, легковых и спецтехники
        for row in ws.iter_rows(max_row=10, values_only=True):
            if not row or not row[0]:
                continue

            # Ищем заголовок "Грузовой"
            if "Грузовой" in str(row[0]):
                # Типы колес для грузовых находятся в ячейках B2-H2
                truck_row = list(ws.iter_rows(min_row=2, max_row=2, min_col=2, max_col=8, values_only=True))[0]
                wheel_types["Грузовой"] = [str(cell).strip() for cell in truck_row if
                                           cell and str(cell).strip() and str(cell).strip() != "None"]

            # Ищем заголовок "Легковой"
            if "Легковой" in str(row[0]):
                # Типы колес для легковых находятся в ячейках J2-R2
                car_row = list(ws.iter_rows(min_row=2, max_row=2, min_col=10, max_col=18, values_only=True))[0]
                wheel_types["Легковой"] = [str(cell).strip() for cell in car_row if
                                           cell and str(cell).strip() and str(cell).strip() != "None"]

            # Ищем заголовок "Спецтехника"
            if "Спецтехника" in str(row[0]):
                # Типы колес для спецтехники находятся в ячейках T2-W2
                special_row = list(ws.iter_rows(min_row=2, max_row=2, min_col=20, max_col=23, values_only=True))[0]
                wheel_types["Спецтехника"] = [str(cell).strip() for cell in special_row if
                                              cell and str(cell).strip() and str(cell).strip() != "None"]

    except Exception as e:
        print(f"Ошибка загрузки типов колес: {e}")
        # Возвращаем значения по умолчанию если не удалось загрузить
        wheel_types = {
            "Грузовой": ["Газели и малые груз.", "16 С Бычок, фотон", "16 С кольцом", "17.5", "19-22,5",
                         "20 с кольцом", "Вездеходы"],
            "Легковой": ["R12", "R13", "R14", "R15", "R16", "R17-18", "R19-20", "R21-23", "R24-26"],
            "Спецтехника": ["16/70-20 12.5/80-18", "16.9-24 16.9-28", "17.5-25 20.5-25", "23.5-25"]
            # Значения по умолчанию
        }

        # Убедимся, что списки не пустые
    if not wheel_types["Грузовой"]:
        wheel_types["Грузовой"] = ["Газели и малые груз.", "16 С Бычок, фотон", "16 С кольцом", "17.5", "19-22,5",
                                   "20 с кольцом", "Вездеходы"]
    if not wheel_types["Легковой"]:
        wheel_types["Легковой"] = ["R12", "R13", "R14", "R15", "R16", "R17-18", "R19-20", "R21-23", "R24-26"]
    if not wheel_types["Спецтехника"]:
        wheel_types["Спецтехника"] = ["16/70-20 12.5/80-18", "16.9-24 16.9-28", "17.5-25 20.5-25", "23.5-25"]

    return wheel_types


# Загружаем типы колес при старте
WHEEL_TYPES = load_wheel_types()


# === Цены услуг и расходников ===
def _parse_price_value(v):
    if isinstance(v, str) and "/" in v:
        parts = [p.strip() for p in v.split("/") if p.strip()]
        if len(parts) == 2:
            try:
                return int(parts[0]), int(parts[1])
            except Exception:
                return 0
    try:
        # Проверяем, является ли значение числом
        if isinstance(v, (int, float)):
            return int(v)
        # Пробуем преобразовать строку в число
        if isinstance(v, str):
            # Убираем пробелы и нецифровые символы
            cleaned = ''.join(c for c in v if c.isdigit())
            if cleaned:
                return int(cleaned)
        return 0
    except Exception:
        return 0


def load_price_table(use_nds: bool = False):
    """Загружает таблицу цен, с НДС или без"""
    price_file = PRICE_NDS_XLSX if use_nds else PRICE_XLSX
    price = {"Легковой": {}, "Грузовой": {}, "Спецтехника": {}}  # Добавляем категорию

    if not price_file.exists():
        return price

    try:
        wb = load_workbook(price_file, data_only=True)
        ws = wb.active

        # Загружаем типы колес
        wheel_types = load_wheel_types()
        truck_wheels = wheel_types["Грузовой"]
        car_wheels = wheel_types["Легковой"]
        special_wheels = wheel_types["Спецтехника"]  # Типы для спецтехники

        # Проходим по всем строкам таблицы для поиска услуг
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            service_name = str(row[0]).strip() if row and row[0] else ""
            if not service_name or service_name in ["Грузовой", "Легковой", "Спецтехника", "Тип колёс", None, "None"]:
                continue

            # Цены для грузовых (колонки B-H)
            for i, wheel_type in enumerate(truck_wheels):
                col_idx = 2 + i  # B=2, C=3, D=4, E=5, F=6, G=7, H=8
                if col_idx - 1 < len(row):
                    price_value = _parse_price_value(row[col_idx - 1])
                    if price_value:
                        key = f"{service_name}|{wheel_type}"
                        price["Грузовой"][key] = price_value

            # Цены для легковых (колонки J-R)
            for i, wheel_type in enumerate(car_wheels):
                col_idx = 10 + i  # J=10, K=11, L=12, M=13, N=14, O=15, P=16, Q=17, R=18
                if col_idx - 1 < len(row):
                    price_value = _parse_price_value(row[col_idx - 1])
                    if price_value:
                        key = f"{service_name}|{wheel_type}"
                        price["Легковой"][key] = price_value

            # Цены для спецтехники (колонки T-W)
            for i, wheel_type in enumerate(special_wheels):
                col_idx = 20 + i  # T=20, U=21, V=22, W=23
                if col_idx - 1 < len(row):
                    price_value = _parse_price_value(row[col_idx - 1])
                    if price_value:
                        key = f"{service_name}|{wheel_type}"
                        price["Спецтехника"][key] = price_value

    except Exception as e:
        print(f"Ошибка загрузки цен ({'с НДС' if use_nds else 'без НДС'}): {e}")

    return price


def load_consumables_table(use_nds: bool = False):
    """Загружает таблица расходников, с НДС или без"""
    consumables_file = CONSUMABLES_NDS_XLSX if use_nds else CONSUMABLES_XLSX
    data = {}
    categories = []

    if not consumables_file.exists():
        return data, categories

    try:
        wb = load_workbook(consumables_file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return data, categories

        # Считываем заголовки категорий
        header1 = rows[0]  # Первая строка с категориями
        header2 = rows[1]  # Вторая строка с подзаголовками

        # Определяем категории и их пары (холодная/горячая)
        category_pairs = []
        i = 2  # Начинаем с колонки C (индекс 2)
        while i < len(header1):
            cat_name = header1[i]
            if cat_name:
                cat_name = str(cat_name).strip()
                # Добавляем пару (категория, холодная) и (категория, горячая)
                category_pairs.append((cat_name, "холодная"))
                if i + 1 < len(header1):
                    category_pairs.append((cat_name, "горячая"))
                categories.append(cat_name)
            i += 2  # Переходим к следующей паре колонок

        # Обрабатываем строки данных
        for row_idx, row in enumerate(rows[2:], start=3):  # Начинаем с 3-й строки
            kind = row[0]
            name = row[1]
            if not kind or not name:
                continue

            kind = str(kind).strip()
            name = str(name).strip()

            # Инициализируем структуру данных
            data.setdefault(kind, {}).setdefault(name, {})

            # Обрабатываем пары колонок для цен
            for pair_idx, (cat_name, temp) in enumerate(category_pairs):
                # Вычисляем индекс колонки в строке
                col_idx = 2 + pair_idx  # Начинаем с колонки C (индекс 2)

                if col_idx < len(row):
                    price_val = row[col_idx]

                    # Проверяем, есть ли цена
                    if price_val is not None and price_val != "":
                        price = _parse_price_value(price_val)
                        if price:
                            data[kind][name][(cat_name, temp)] = price
                            # Для отладки
                            if kind == "Грибок":
                                print(f"Грибок {name}: {cat_name} {temp} = {price}")

    except Exception as e:
        print(f"Ошибка загрузки расходников ({'с НДС' if use_nds else 'без НДС'}): {e}")
        import traceback
        traceback.print_exc()

    return data, categories


# Загружаем таблицы цен и расходников
PRICE_TABLE = load_price_table()
PRICE_TABLE_NDS = load_price_table(use_nds=True)
CONSUMABLES_TABLE, CONSUMABLE_CATEGORIES = load_consumables_table()
CONSUMABLES_TABLE_NDS, CONSUMABLE_CATEGORIES_NDS = load_consumables_table(use_nds=True)

CONSUMABLE_SERVICE_MAP = {
    "Пластырь №": "Пластырь",
    "Грибок №": "Грибок",
    "Удлинитель": "Удлинитель",
    "Грузики": "Грузики",
}

SERVICE_PRICE_NAME = {
    "Снятие/установка": "Снятие, установка наружное/внутреннее",
    "Вентиль легковой": "Вентиль легковой (хром/черный)",
    "Пластырь №": "Пластырь",
    "Грибок №": "Грибок",
    "Удлинитель": "Удлинитель ",
}


# === Чек и текст суммы ===
def ruble_suffix(n: int) -> str:
    n_abs = abs(n) % 100
    n1 = n_abs % 10
    if 11 <= n_abs <= 19:
        return "рублей"
    if n1 == 1:
        return "рубль"
    if 2 <= n1 <= 4:
        return "рубля"
    return "рублей"


def make_total_text(total: int) -> str:
    words = num2words(total, lang='ru').capitalize()
    return f"{words} {ruble_suffix(total)}"


# === Экспорт PDF ===
def export_pdf_via_excel(xlsx_path: Path, pdf_path: Path, a5: bool = True, landscape: bool = False) -> bool:
    try:
        import win32com.client as win32
        from win32com.client import constants
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(str(xlsx_path.resolve()))
        ws = wb.Worksheets(1)
        if a5:
            ws.PageSetup.PaperSize = constants.xlPaperA5
        ws.PageSetup.Orientation = constants.xlLandscape if landscape else constants.xlPortrait
        xlTypePDF = 0
        wb.ExportAsFixedFormat(xlTypePDF, str(pdf_path.resolve()))
        wb.Close(SaveChanges=False)
        excel.Quit()
        return True
    except Exception:
        return False


def export_pdf_via_libreoffice(xlsx_path: Path, pdf_path: Path) -> bool:
    try:
        outdir = pdf_path.parent
        cmd = ["soffice", "--headless", "--convert-to", "pdf", "--outdir", str(outdir), str(xlsx_path.resolve())]
        subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        produced = outdir / (xlsx_path.stem + ".pdf")
        if produced.exists():
            if produced != pdf_path:
                produced.replace(pdf_path)
            return True
        return False
    except Exception:
        return False


def _write_to_excel(ws, data: dict) -> int:
    ws[CELL_CUSTOMER] = data["customer_display"]
    plate_text = data.get("plate", "")
    trailer = data.get("trailer", "")
    if trailer and trailer != "Без прицепа":
        plate_text = f"{plate_text}, {trailer}" if plate_text else trailer
    ws[CELL_PLATE] = plate_text
    ws[CELL_DRIVER] = data["driver_name"]
    defect_value = data["defect"]
    # ИЗМЕНЕНО: если дефект "Пропустить", оставляем поле пустым
    ws[CELL_DEFECT_LINE1] = "" if defect_value == "Пропустить" else defect_value
    ws[CELL_DEFECT_LINE2] = ""

    ws[CELL_ISSUED_TO] = data["issued_to"]
    ws[CELL_DATE] = dt.now().strftime("%d.%m.%Y")  # Исправляем здесь
    # ИЗМЕНЕНО: убрано заполнение фамилии механика (оставляем пустым)
    ws[CELL_MECHANIC] = ""
    total = 0

    # Записываем услуги
    for idx, service_name in enumerate(SERVICES):
        row = SERVICES_START_ROW + idx
        detail = data["services"].get(service_name, {})
        qty = detail.get("qty", 0)
        price = detail.get("price", 0)
        cost = detail.get("cost", qty * price)
        ws[f"{COL_QTY}{row}"] = qty if qty else ""
        ws[f"{COL_PRICE}{row}"] = price if qty else ""
        ws[f"{COL_COST}{row}"] = cost if qty else ""
        total += cost

    # ИЗМЕНЕНО: Записываем колеса в ячейки BF38:BQ38, BF39:BQ39, BF40:BQ40
    wheels = data.get("wheels", [])
    for i, wheel_entry in enumerate(wheels[:3]):  # Максимум 3 строки для колес
        row = WHEELS_START_ROW + i
        try:
            # Объединяем ячейки BF- BQ для текущей строки
            start_col = "BF"
            end_col = "BQ"
            ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
            # Записываем значение
            ws[f"{start_col}{row}"] = wheel_entry
            # Применяем выравнивание по центру по горизонтали и вертикали
            ws[f"{start_col}{row}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        except Exception as e:
            print(f"Ошибка записи колеса {i + 1}: {e}")

    # ИЗМЕНЕНО: Итоговая сумма в BR41:CO41 с выравниванием по центру
    # Сначала объединяем ячейки BR41:CO41
    ws.merge_cells(f"BR41:CO41")
    # Записываем значение в объединенную ячейку
    ws["BR41"] = total
    # Применяем выравнивание по центру по горизонтали и вертикали
    ws["BR41"].alignment = Alignment(horizontal='center', vertical='center')

    # ИЗМЕНЕНО: Сумма прописью в A43:BE43
    ws[CELL_TOTAL_TEXT] = make_total_text(total)
    return total


# === Обновляем функции заполнения Excel ===
def fill_excel_only(data: dict) -> Path:
    if not TEMPLATE_XLSX.exists():
        raise FileNotFoundError(f"Не найден шаблон: {TEMPLATE_XLSX}")

    # Получаем текущую папку дня
    day_folder = get_current_day_folder()

    # Формируем имя файла с временем
    current_time = dt.now()
    dt_str = current_time.strftime("%Y%m%d_%H%M%S")
    xlsx_out = day_folder / f"наряд_{dt_str}.xlsx"

    wb = load_workbook(TEMPLATE_XLSX)
    ws = wb.active
    _write_to_excel(ws, data)
    wb.save(xlsx_out)
    return xlsx_out


def fill_excel_and_export_pdf(data: dict) -> tuple[Path, Path]:
    if not TEMPLATE_XLSX.exists():
        raise FileNotFoundError(f"Не найден шаблон: {TEMPLATE_XLSX}")

    # Получаем текущую папку дня
    day_folder = get_current_day_folder()

    # Формируем имя файла с временем
    current_time = dt.now()
    dt_str = current_time.strftime("%Y%m%d_%H%M%S")
    xlsx_out = day_folder / f"наряд_{dt_str}.xlsx"
    pdf_out = day_folder / f"наряд_{dt_str}.pdf"

    wb = load_workbook(TEMPLATE_XLSX)
    ws = wb.active
    _write_to_excel(ws, data)
    wb.save(xlsx_out)

    ok = export_pdf_via_excel(xlsx_out, pdf_out, a5=True, landscape=False)
    if not ok and not export_pdf_via_libreoffice(xlsx_out, pdf_out):
        raise RuntimeError(
            "Не удалось экспортировать в PDF. Проверьте наличие Microsoft Excel (или LibreOffice в PATH).")
    return xlsx_out, pdf_out

def check_and_create_day_folders():
    """
    Проверяет и создает папки при запуске программы.
    Эта функция запускается при старте.
    """
    try:
        # Просто получаем текущую папку дня - она создастся автоматически
        current_folder = get_current_day_folder()
        print(f"✅ Текущая папка дня: {current_folder}")

        # Проверяем, есть ли папки за последние 7 дней
        for i in range(7):
            check_date = dt.utcnow() - datetime.timedelta(days=i)
            # МСК = UTC+3
            moscow_time = check_date + datetime.timedelta(hours=3)

            # Для каждой даты проверяем, была ли она после 07:00
            if moscow_time.hour < 7:
                folder_date = moscow_time - datetime.timedelta(days=1)
            else:
                folder_date = moscow_time

            folder_name = folder_date.strftime("%Y-%m-%d")
            day_folder = OUTPUT_BASE_DIR / folder_name

            if not day_folder.exists():
                day_folder.mkdir(exist_ok=True, parents=True)
                print(f"📁 Создана папка для даты: {folder_name}")

    except Exception as e:
        print(f"❌ Ошибка при создании папок: {e}")

# === Скролл-фреймы ===
class VScrollFrame(ttk.Frame):
    def __init__(self, master, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.canvas = tk.Canvas(self, highlightthickness=0)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.vsb.grid(row=0, column=1, sticky="ns")

        self.inner = ttk.Frame(self.canvas)
        self.inner_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self._need_scroll = False

        def _update_scrollregion(event=None):
            self.canvas.itemconfig(self.inner_id, width=self.canvas.winfo_width())
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
            need = (self.inner.winfo_reqheight() > self.canvas.winfo_height())
            if need != self._need_scroll:
                self._need_scroll = need
                if self._need_scroll:
                    self.vsb.grid()
                else:
                    self.vsb.grid_remove()
                    self.canvas.yview_moveto(0)

        self.inner.bind("<Configure>", _update_scrollregion)
        self.canvas.bind("<Configure>", _update_scrollregion)

        # колёсико по наведению
        def _bind_wheel(_=None):
            if self._need_scroll:
                self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        def _unbind_wheel(_=None):
            self.canvas.unbind_all("<MouseWheel>")

        for w in (self.canvas, self.inner):
            w.bind("<Enter>", _bind_wheel)
            w.bind("<Leave>", _unbind_wheel)

    def _on_mousewheel(self, event):
        if not self._need_scroll:
            return
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


class HighlightList(tb.Frame):
    def __init__(self, master, on_select, keybind_parent=None):
        super().__init__(master)
        self.on_select = on_select
        self.items = []
        self.current_index = 0
        self.visible = True
        self.keybind_parent = keybind_parent or master
        self._bind_ids = []

        self.grid_columnconfigure(0, weight=1)

        self.canvas = tk.Canvas(self, highlightthickness=0, height=160)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.vsb.grid(row=0, column=1, sticky="ns")

        self.inner = tb.Frame(self.canvas)
        self.inner_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self._need_scroll = False

        def _update(event=None):
            self.canvas.itemconfig(self.inner_id, width=self.canvas.winfo_width())
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
            need = (self.inner.winfo_reqheight() > self.canvas.winfo_height())
            if need != self._need_scroll:
                self._need_scroll = need
                if need:
                    self.vsb.grid()
                else:
                    self.vsb.grid_remove()
                    self.canvas.yview_moveto(0)

        self.inner.bind("<Configure>", _update)
        self.canvas.bind("<Configure>", _update)

        # колесо по наведению
        def _bind_wheel(_=None):
            if self._need_scroll:
                self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        def _unbind_wheel(_=None):
            self.canvas.unbind_all("<MouseWheel>")

        for w in (self.canvas, self.inner):
            w.bind("<Enter>", _bind_wheel)
            w.bind("<Leave>", _unbind_wheel)

        self._bind_ids.append(self.keybind_parent.bind("<Up>", self._move_up))
        self._bind_ids.append(self.keybind_parent.bind("<Down>", self._move_down))
        self._bind_ids.append(self.keybind_parent.bind("<Return>", self._enter))

    def destroy(self):
        for bid in self._bind_ids:
            try:
                self.keybind_parent.unbind("<Up>", bid)
                self.keybind_parent.unbind("<Down>", bid)
                self.keybind_parent.unbind("<Return>", bid)
            except Exception:
                pass
        super().destroy()

    def show(self):
        self.grid()
        self.visible = True

    def hide(self):
        self.grid_remove()
        self.visible = False

    def set_items(self, names, query):
        for _, row in self.items:
            row.destroy()
        self.items.clear()

        q = (query or "").lower().strip()

        def highlight_text(name: str):
            if not q:
                return name, None, None
            i = name.lower().find(q)
            if i >= 0:
                return name, i, len(q)
            return name, None, None

        for idx, name in enumerate(names):
            text, start, ln = highlight_text(name)
            row = tb.Frame(self.inner)
            row.pack(fill=X, padx=4, pady=2)

            pre = text[:start] if start is not None else text
            match = text[start:start + ln] if start is not None else ""
            post = text[start + ln:] if start is not None else ""

            tb.Label(row, text=pre, anchor="w").pack(side=LEFT)
            if match:
                tb.Label(row, text=match, bootstyle="warning").pack(side=LEFT)
            if post:
                tb.Label(row, text=post, anchor="w").pack(side=LEFT)

            def _click_factory(n=name):
                return lambda e: self.on_select(n)

            row.bind("<Button-1>", _click_factory())
            for child in row.winfo_children():
                child.bind("<Button-1>", _click_factory())

            self.items.append((name, row))

        self.current_index = 0
        self._refresh_active_row()

        if names:
            self.show()
        else:
            self.hide()

    def _refresh_active_row(self):
        for i, (_, row) in enumerate(self.items):
            row.configure(bootstyle=("info" if i == self.current_index else "secondary"))

    def _move_up(self, event=None):
        if not self.visible or not self.items: return
        self.current_index = (self.current_index - 1) % len(self.items)
        self._refresh_active_row()

    def _move_down(self, event=None):
        if not self.visible or not self.items: return
        self.current_index = (self.current_index + 1) % len(self.items)
        self._refresh_active_row()

    def _enter(self, event=None):
        if not self.visible or not self.items: return
        name, _ = self.items[self.current_index]
        self.on_select(name)


class ConsumableDialog(tb.Toplevel):
    def __init__(self, parent, kind: str, qty: int, use_nds: bool = False):
        super().__init__(parent)
        self.title(kind)
        self.result = None
        self.grab_set()

        # Используем правильную таблицу расходников в зависимости от НДС
        consumables_table = CONSUMABLES_TABLE_NDS if use_nds else CONSUMABLES_TABLE
        consumable_categories = CONSUMABLE_CATEGORIES_NDS if use_nds else CONSUMABLE_CATEGORIES

        names = sorted(consumables_table.get(kind, {}).keys())

        # ИЗМЕНЕНО: для Грибка показываем только выбор названия
        if kind == "Грибок":
            self.vars = []
            for i in range(qty):
                row = tb.Frame(self, padding=4)
                row.grid(row=i, column=0, sticky="we")
                name_var = tk.StringVar(value=(names[0] if names else ""))
                tb.Label(row, text="Название:").pack(side=LEFT, padx=4)
                tb.Combobox(row, values=names, textvariable=name_var, state="readonly", width=30).pack(side=LEFT,
                                                                                                       padx=4)
                self.vars.append((name_var,))
        else:
            cats = consumable_categories
            temps = ["холодная", "горячая"]
            self.vars = []
            for i in range(qty):
                row = tb.Frame(self, padding=4)
                row.grid(row=i, column=0, sticky="we")
                name_var = tk.StringVar(value=(names[0] if names else ""))
                cat_var = tk.StringVar(value=(cats[0] if cats else ""))
                temp_var = tk.StringVar(value=temps[0])
                tb.Label(row, text="Название:").pack(side=LEFT, padx=2)
                tb.Combobox(row, values=names, textvariable=name_var, state="readonly", width=20).pack(side=LEFT,
                                                                                                       padx=2)
                tb.Label(row, text="Категория:").pack(side=LEFT, padx=2)
                tb.Combobox(row, values=cats, textvariable=cat_var, state="readonly", width=20).pack(side=LEFT, padx=2)
                tb.Label(row, text="Температура:").pack(side=LEFT, padx=2)
                tb.Combobox(row, values=temps, textvariable=temp_var, state="readonly", width=12).pack(side=LEFT,
                                                                                                       padx=2)
                self.vars.append((name_var, cat_var, temp_var))

        btn = tb.Button(self, text="OK", command=self._ok)
        btn.grid(row=qty, column=0, pady=6)

    def _ok(self):
        res = []
        if len(self.vars[0]) == 1:  # Для Грибка
            for (n,) in self.vars:
                # Для Грибка используем первую категорию и холодную температуру по умолчанию
                res.append((n.get(), "Грузовые автомобили 230-445 мм", "холодная"))
        else:  # Для остальных расходников
            for n, c, t in self.vars:
                res.append((n.get(), c.get(), t.get()))
        self.result = res
        self.destroy()


class SimpleConsumableDialog(tb.Toplevel):
    """Упрощенный диалог для выбора только названия (для Грибка)"""

    def __init__(self, parent, kind: str, qty: int, use_nds: bool = False):
        super().__init__(parent)
        self.title(kind)
        self.result = None
        self.grab_set()

        # Используем правильную таблицу расходников в зависимости от НДС
        consumables_table = CONSUMABLES_TABLE_NDS if use_nds else CONSUMABLES_TABLE

        names = sorted(consumables_table.get(kind, {}).keys())

        self.vars = []
        for i in range(qty):
            row = tb.Frame(self, padding=4)
            row.grid(row=i, column=0, sticky="we")
            name_var = tk.StringVar(value=(names[0] if names else ""))
            tb.Label(row, text="Название:").pack(side=LEFT, padx=4)
            tb.Combobox(row, values=names, textvariable=name_var, state="readonly", width=30).pack(side=LEFT, padx=4)
            self.vars.append((name_var,))

        btn = tb.Button(self, text="OK", command=self._ok)
        btn.grid(row=qty, column=0, pady=6)

    def _ok(self):
        res = []
        for (n,) in self.vars:
            # Для Грибка используем фиксированные значения категории и температуры
            res.append((n.get(), "Грузовые автомобили 230-445 мм", "холодная"))
        self.result = res
        self.destroy()


# === Приложение ===
class WorkOrderApp:
    def __init__(self, root: tb.Window):
        self.root = root
        self.root.title("Наряд-Заказ — v3.0")
        self.root.geometry("1280x900")

        # Проверяем и создаем папки при запуске
        check_and_create_day_folders()

        # Верхняя панель
        topbar = tb.Frame(self.root, padding=10)
        tb.Label(topbar, text="Наряд‑Заказ", font=("-size", 18, "-weight", "bold")).pack(side=LEFT)

        # Добавляем информацию о текущей папке
        current_folder = get_current_day_folder()
        self.folder_label = tb.Label(
            topbar,
            text=f"Папка: {current_folder.name}",
            font=("-size", 11),
            bootstyle="info"
        )
        self.folder_label.pack(side=LEFT, padx=15)

        # ДОБАВЛЯЕМ ЗДЕСЬ НОВУЮ КНОПКУ (после кнопки "Проверить синхронизацию")
        tb.Button(topbar, text="💰 Посчитать день", bootstyle="info",
                  command=self.calculate_day_total).pack(side=RIGHT, padx=6)

        # Существующие кнопки
        tb.Button(topbar, text="Создать наряд", bootstyle="primary", command=self.open_create_form).pack(side=RIGHT,
                                                                                                         padx=6)
        tb.Button(topbar, text="Админ‑панель", bootstyle="secondary", command=self.open_admin_panel).pack(side=RIGHT,
                                                                                                          padx=6)
        tb.Button(topbar, text="Обновить списки", bootstyle="warning", command=self.refresh_lists).pack(side=RIGHT,
                                                                                                        padx=6)

        # Кнопка проверки синхронизации
        tb.Button(topbar, text="Проверить синхронизацию", bootstyle="info", command=self.check_sync_status).pack(
            side=RIGHT, padx=6)

        # Кнопка открытия текущей папки
        tb.Button(topbar, text="📁 Открыть папку дня", bootstyle="success",
                  command=lambda: self.open_current_folder()).pack(side=RIGHT, padx=6)

        topbar.pack(fill=X)

        self.root.bind("<Control-n>", lambda e: self.open_create_form())

        # Плейсхолдер
        self.placeholder = tb.Frame(self.root, padding=20)
        tb.Label(self.placeholder, text="Нажмите «Создать наряд» или Ctrl+N", bootstyle="secondary", font=("-size", 14)).pack()
        self.placeholder.pack(fill=BOTH, expand=True)

        self._create_form_window = None  # ссылка, чтобы обновлять виджеты после админки
        self.current_page = 0  # 0 = Page 1, 1 = Page 2

        # Запускаем периодическое обновление информации о папке
        self.update_folder_info()

    def setup_fonts(self):
        """Настраивает размеры шрифтов для виджетов"""
        # Увеличиваем размер шрифта по умолчанию
        default_font = ("TkDefaultFont", 11)  # Было примерно 9
        self.root.option_add("*Font", default_font)

        # Настройка для разных виджетов
        label_font = ("TkDefaultFont", 11)
        entry_font = ("TkDefaultFont", 11)
        button_font = ("TkDefaultFont", 11)
        combobox_font = ("TkDefaultFont", 11)

        # Применяем настройки
        self.root.option_add("*Label*Font", label_font)
        self.root.option_add("*Entry*Font", entry_font)
        self.root.option_add("*Button*Font", button_font)
        self.root.option_add("*Combobox*Font", combobox_font)
        self.root.option_add("*Checkbutton*Font", label_font)
        self.root.option_add("*Radiobutton*Font", label_font)
        self.root.option_add("*Listbox*Font", entry_font)
        self.root.option_add("*Text*Font", entry_font)

    def update_folder_info(self):
        """Обновляет информацию о текущей папке в интерфейсе"""
        try:
            current_folder = get_current_day_folder()
            self.folder_label.config(text=f"Папка: {current_folder.name}")

            # Подсчитываем количество файлов в папке
            file_count = len(list(current_folder.glob("*")))
            self.folder_label.config(
                text=f"Папка: {current_folder.name} (файлов: {file_count})"
            )
        except Exception as e:
            print(f"Ошибка обновления информации о папке: {e}")

        # Обновляем каждые 5 минут
        self.root.after(300000, self.update_folder_info)  # 5 минут = 300000 мс

    def open_current_folder(self):
        """Открывает текущую папку дня в проводнике"""
        try:
            current_folder = get_current_day_folder()
            os.startfile(str(current_folder.resolve()))
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть папку:\n{e}", parent=self.root)

    def check_sync_status(self):
        """Проверяет статус синхронизации с Google Sheets"""
        try:
            from google.oauth2.service_account import Credentials
            import gspread

            if not GOOGLE_CREDENTIALS_FILE.exists():
                messagebox.showwarning("Синхронизация",
                                       "Файл учетных данных не найден.\n\n"
                                       "Для работы онлайн-синхронизации необходимо:\n"
                                       "1. Создать проект в Google Cloud Console\n"
                                       "2. Включить Google Sheets API\n"
                                       "3. Создать сервисный аккаунт\n"
                                       "4. Скачать credentials.json в папку с программой\n"
                                       "5. Поделиться Google таблицей с email сервисного аккаунта",
                                       parent=self.root)
                return

            # Пробуем подключиться к Google Sheets
            df = read_companies_from_google()
            if df is not None:
                messagebox.showinfo("Синхронизация",
                                    "✅ Синхронизация с Google Sheets работает!\n"
                                    f"Загружено {len(df)} компаний.",
                                    parent=self.root)
            else:
                messagebox.showwarning("Синхронизация",
                                       "❌ Не удалось подключиться к Google Sheets.\n"
                                       "Проверьте настройки и подключение к интернету.",
                                       parent=self.root)

        except ImportError:
            messagebox.showerror("Синхронизация",
                                 "Библиотеки для работы с Google Sheets не установлены.\n\n"
                                 "Установите их командой:\n"
                                 "pip install gspread google-auth",
                                 parent=self.root)

    def calculate_day_total(self):
        """Подсчитывает общую стоимость всех нарядов за текущий день"""
        try:
            # Получаем текущую папку дня
            current_folder = get_current_day_folder()

            # Ищем все Excel файлы в папке дня
            excel_files = list(current_folder.glob("наряд_*.xlsx"))

            if not excel_files:
                messagebox.showinfo(
                    "Подсчет дня",
                    f"В папке дня '{current_folder.name}' нет наряд-заказов.",
                    parent=self.root
                )
                return

            total_sum = 0
            processed_files = 0
            detailed_info = []

            # Открываем и анализируем каждый файл
            for excel_file in excel_files:
                try:
                    wb = load_workbook(excel_file, data_only=True)
                    ws = wb.active

                    # Читаем сумму из ячейки BR41 (итоговая сумма)
                    # ВАЖНО: Ячейка BR41:CO41 объединена, читаем из BR41
                    total_cell = ws["BR41"].value

                    if total_cell:
                        try:
                            # Пробуем преобразовать значение в число
                            if isinstance(total_cell, str):
                                # Убираем пробелы и лишние символы
                                total_cell = ''.join(c for c in total_cell if c.isdigit() or c == '.')
                            amount = float(total_cell)
                            total_sum += amount
                            processed_files += 1

                            # Читаем дополнительную информацию для детализации
                            company = ws[CELL_CUSTOMER].value or "Не указано"
                            plate = ws[CELL_PLATE].value or "Не указано"
                            date = ws[CELL_DATE].value or "Не указано"

                            detailed_info.append(
                                f"• {excel_file.name}:\n"
                                f"  Компания: {company}\n"
                                f"  Номер: {plate}\n"
                                f"  Дата: {date}\n"
                                f"  Сумма: {amount:,.2f} руб.\n"
                            )
                        except (ValueError, TypeError) as e:
                            print(f"Ошибка чтения суммы из {excel_file.name}: {e}")
                            detailed_info.append(f"• {excel_file.name}: ОШИБКА чтения суммы\n")
                    else:
                        detailed_info.append(f"• {excel_file.name}: НЕТ суммы\n")

                except Exception as e:
                    print(f"Ошибка обработки файла {excel_file.name}: {e}")
                    detailed_info.append(f"• {excel_file.name}: ОШИБКА обработки\n")

            # Формируем итоговое сообщение
            message_text = (
                f"📊 ОТЧЕТ ЗА ДЕНЬ\n"
                f"Папка: {current_folder.name}\n"
                f"Всего файлов: {len(excel_files)}\n"
                f"Обработано успешно: {processed_files}\n"
                f"Общая сумма: {total_sum:,.2f} руб.\n"
            )

            # Добавляем детализацию, если есть успешно обработанные файлы
            if processed_files > 0:
                message_text += f"\n📋 Детализация:\n" + "\n".join(detailed_info)

            # Показываем результат в отдельном окне с прокруткой
            self._show_day_report(message_text, current_folder.name, total_sum, len(excel_files), current_folder)

        except Exception as e:
            messagebox.showerror(
                "Ошибка подсчета",
                f"Не удалось подсчитать суммы:\n{str(e)}",
                parent=self.root
            )

    def _show_day_report(self, report_text, folder_name, total_sum, file_count, current_folder):
        """Показывает детализированный отчет в отдельном окне"""
        report_window = tb.Toplevel(self.root)
        report_window.title(f"Отчет за день: {folder_name}")
        report_window.geometry("800x600")
        report_window.resizable(True, True)

        # Создаем фрейм с прокруткой
        main_frame = tb.Frame(report_window, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Заголовок
        header_frame = tb.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))

        tb.Label(
            header_frame,
            text=f"📊 ИТОГИ ДНЯ",
            font=("-size", 16, "-weight", "bold"),
            bootstyle="info"
        ).pack()

        tb.Label(
            header_frame,
            text=f"Папка: {folder_name} | Файлов: {file_count}",
            font=("-size", 11)
        ).pack()

        # Общая сумма крупным шрифтом
        total_frame = tb.Frame(main_frame)
        total_frame.pack(fill=tk.X, pady=10)

        tb.Label(
            total_frame,
            text=f"ОБЩАЯ СУММА:",
            font=("-size", 12)
        ).pack(side=tk.LEFT)

        tb.Label(
            total_frame,
            text=f"{total_sum:,.2f} руб.",
            font=("-size", 14, "-weight", "bold"),
            bootstyle="success"
        ).pack(side=tk.LEFT, padx=(10, 0))

        # Детализация с прокруткой
        detail_label = tb.Label(main_frame, text="Детализация нарядов:")
        detail_label.pack(anchor=tk.W, pady=(10, 5))

        # Создаем текстовое поле с прокруткой для детализации
        text_frame = tb.Frame(main_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)

        text_scroll = tb.Scrollbar(text_frame)
        text_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        report_text_widget = tk.Text(
            text_frame,
            wrap=tk.WORD,
            yscrollcommand=text_scroll.set,
            font=("Consolas", 10),
            height=20
        )
        report_text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        text_scroll.config(command=report_text_widget.yview)

        # Вставляем текст отчета
        report_text_widget.insert(tk.END, report_text)
        report_text_widget.config(state=tk.DISABLED)  # Делаем только для чтения

        # Кнопки внизу
        button_frame = tb.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))

        # Кнопка копировать в буфер
        def copy_to_clipboard():
            self.root.clipboard_clear()
            self.root.clipboard_append(f"Отчет за день {folder_name}\n")
            self.root.clipboard_append(f"Всего файлов: {file_count}\n")
            self.root.clipboard_append(f"Общая сумма: {total_sum:,.2f} руб.\n")
            messagebox.showinfo("Скопировано", "Итог скопирован в буфер обмена", parent=report_window)

        tb.Button(
            button_frame,
            text="📋 Копировать итог",
            bootstyle="secondary",
            command=copy_to_clipboard
        ).pack(side=tk.LEFT, padx=5)

        # Кнопка экспорта в текстовый файл
        def export_to_file():
            try:
                export_path = current_folder / f"отчет_{folder_name}.txt"
                with open(export_path, 'w', encoding='utf-8') as f:
                    f.write(f"ОТЧЕТ ЗА ДЕНЬ: {folder_name}\n")
                    f.write("=" * 50 + "\n")
                    f.write(f"Всего наряд-заказов: {file_count}\n")
                    f.write(f"Общая сумма: {total_sum:,.2f} руб.\n")
                    f.write("=" * 50 + "\n")
                    f.write("Детализация:\n")
                    f.write(report_text)

                messagebox.showinfo(
                    "Экспорт завершен",
                    f"Отчет сохранен в файл:\n{export_path}",
                    parent=report_window
                )
                os.startfile(str(export_path.parent))
            except Exception as e:
                messagebox.showerror("Ошибка экспорта", str(e), parent=report_window)

        tb.Button(
            button_frame,
            text="💾 Экспорт в файл",
            bootstyle="info",
            command=export_to_file
        ).pack(side=tk.LEFT, padx=5)

        # Кнопка закрыть
        tb.Button(
            button_frame,
            text="Закрыть",
            bootstyle="primary",
            command=report_window.destroy
        ).pack(side=tk.RIGHT, padx=5)

    def refresh_lists(self):
        reload_companies_globals()
        global WHEEL_TYPES, PRICE_TABLE, PRICE_TABLE_NDS, CONSUMABLES_TABLE, CONSUMABLES_TABLE_NDS, WHEELS_LIST

        WHEEL_TYPES = load_wheel_types()
        PRICE_TABLE = load_price_table()
        PRICE_TABLE_NDS = load_price_table(use_nds=True)
        CONSUMABLES_TABLE, CONSUMABLE_CATEGORIES = load_consumables_table()
        CONSUMABLES_TABLE_NDS, CONSUMABLE_CATEGORIES_NDS = load_consumables_table(use_nds=True)

        # НОВОЕ: Обновляем список колес
        WHEELS_LIST = load_wheels_from_excel()

        # если форма открыта — обновим виджеты
        self._apply_companies_to_form(self._create_form_window)

        # НОВОЕ: Обновляем ComboBox с колесами в форме, если она открыта
        self._update_wheels_in_form()

        messagebox.showinfo("Готово", "Справочник компаний, типы колес и список колес обновлены.", parent=self.root)

    def _update_wheels_in_form(self):
        """Обновляет список колес в открытой форме"""
        if not hasattr(self, "wheel_combo") or not self._widget_exists(self.wheel_combo):
            return

        # Обновляем значения в ComboBox
        self.wheel_combo["values"] = WHEELS_LIST

        # Устанавливаем первое значение, если список не пустой
        if WHEELS_LIST:
            self.wheel_combo.set(WHEELS_LIST[0])
        else:
            self.wheel_combo.set("")

    # ===== Создание наряда =====
    def open_create_form(self):
        win = tb.Toplevel(self.root)
        self._create_form_window = win
        win.title("Создать наряд")
        win.geometry("1100x850")
        win.resizable(True, True)
        win.option_add("*Font", ("TkDefaultFont", 11))
        try:
            win.state('zoomed')
        except Exception:
            pass

        # хоткеи формы
        win.bind("<Control-s>", lambda e: self._build_xlsx_only())
        win.bind("<Escape>", lambda e: win.destroy())
        self._form_parent = win

        # Контейнер для страниц
        self.container = tb.Frame(win)
        self.container.pack(fill=BOTH, expand=True, padx=8, pady=8)
        self.container.grid_rowconfigure(0, weight=1)
        self.container.grid_columnconfigure(0, weight=1)

        # Создаем обе страницы
        self.page1 = tb.Frame(self.container)
        self.page2 = tb.Frame(self.container)

        # Панель навигации (внизу)
        self.nav_frame = tb.Frame(win)
        self.nav_frame.pack(fill=X, padx=8, pady=(0, 8))

        self.btn_back = tb.Button(self.nav_frame, text="← Назад", bootstyle="secondary",
                                  command=self._go_to_page1, state=DISABLED)
        self.btn_back.pack(side=LEFT, padx=4)

        self.page_label = tb.Label(self.nav_frame, text="Страница 1/2")
        self.page_label.pack(side=LEFT, expand=True)

        self.btn_next = tb.Button(self.nav_frame, text="Далее →", bootstyle="primary",
                                  command=self._go_to_page2)
        self.btn_next.pack(side=RIGHT, padx=4)

        self.btn_create = tb.Button(self.nav_frame, text="Сформировать Excel (Ctrl+S)", bootstyle="success",
                                    command=self._build_xlsx_only)
        self.btn_create.pack(side=RIGHT, padx=4)
        self.btn_create.pack_forget()  # Сначала скрываем

        # Инициализация страниц
        self._init_page1()
        self._init_page2()

        # Показываем первую страницу
        self.page1.grid(row=0, column=0, sticky="nsew")
        self.current_page = 0

        # Корректное отключение trace/биндов при закрытии окна
        def _cleanup():
            try:
                if hasattr(self, "company_query"):
                    self.company_query.trace_remove("write", self._company_query_trace)
            except Exception:
                pass
            try:
                if hasattr(self, "search_results"):
                    self.search_results.destroy()
            except Exception:
                pass
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", _cleanup)

    def _init_page1(self):
        """Инициализация страницы 1: Клиентские данные"""
        page1 = self.page1
        page1.grid_columnconfigure(0, weight=1)

        # ===== Левая колонка =====
        pad = {'padx': 8, 'pady': 4}  # Уменьшены вертикальные отступы

        # Заказчик
        frm_customer = tb.Labelframe(page1, text="Заказчик", padding=10)  # Уменьшен padding
        frm_customer.grid(row=0, column=0, sticky="we", **pad)
        frm_customer.grid_columnconfigure(1, weight=1)

        self.customer_type = tk.StringVar(value="Частное лицо")
        tb.Radiobutton(frm_customer, text="Частное лицо", variable=self.customer_type, value="Частное лицо",
                       command=self._on_customer_type_changed).grid(row=0, column=0, sticky=NW, padx=4,
                                                                    pady=2)  # Уменьшен pady
        tb.Radiobutton(frm_customer, text="Компания", variable=self.customer_type, value="Компания",
                       command=self._on_customer_type_changed).grid(row=0, column=1, sticky=NW, padx=4, pady=2)

        tb.Label(frm_customer, text="Поиск компании или номера (Ctrl+F):").grid(row=1, column=0, sticky=NW, padx=4,
                                                                                pady=2)

        self.company_query = tk.StringVar(value="")
        self.entry_company_query = tb.Entry(frm_customer, textvariable=self.company_query)
        self.entry_company_query.grid(row=1, column=1, sticky="we", padx=4, pady=2)

        def focus_search(event=None):
            self.entry_company_query.focus_set()
            self.entry_company_query.selection_range(0, tk.END)

        self._form_parent.bind("<Control-f>", focus_search)

        def on_pick_company(name):
            self.company_selected.set(name)
            self._update_company_meta()

        self.search_results = HighlightList(frm_customer, on_select=on_pick_company, keybind_parent=self._form_parent)
        self.search_results.grid(row=2, column=0, columnspan=2, sticky="we", padx=2, pady=(0, 4))
        self.search_results.canvas.configure(height=120)  # Уменьшена высота списка поиска

        tb.Label(frm_customer, text="Компания:").grid(row=3, column=0, sticky=NW, padx=4, pady=2)
        self.company_selected = tk.StringVar(value=(ALL_COMPANY_NAMES[0] if ALL_COMPANY_NAMES else ""))
        self.cmb_company = tb.Combobox(frm_customer, textvariable=self.company_selected, values=ALL_COMPANY_NAMES,
                                       state="readonly")
        self.cmb_company.grid(row=3, column=1, sticky="we", padx=4, pady=2)

        tb.Label(frm_customer, text="ИНН:").grid(row=4, column=0, sticky=NW, padx=4, pady=2)
        self.company_inn_var = tk.StringVar(value="")
        tb.Label(frm_customer, textvariable=self.company_inn_var, bootstyle="secondary").grid(row=4, column=1,
                                                                                              sticky="w", padx=4,
                                                                                              pady=2)

        tb.Label(frm_customer, text="НДС:").grid(row=5, column=0, sticky=NW, padx=4, pady=2)
        self.company_vat_var = tk.StringVar(value="")
        tb.Label(frm_customer, textvariable=self.company_vat_var, bootstyle="secondary").grid(row=5, column=1,
                                                                                              sticky="w", padx=4,
                                                                                              pady=2)

        def apply_filter(*_):
            q = self.company_query.get()
            values = filter_companies(q)
            self.cmb_company["values"] = values
            if values:
                self.cmb_company.set(values[0])
            else:
                self.cmb_company.set("")
            self.search_results.set_items(values[:50], q.strip().lower())
            self._update_company_meta()

        self._company_query_trace = self.company_query.trace_add("write", apply_filter)
        self.cmb_company.bind("<<ComboboxSelected>>", lambda e: self._update_company_meta())
        apply_filter()

        # Госномер
        frm_plate = tb.Labelframe(page1, text="Гос. номер", padding=10)
        frm_plate.grid(row=1, column=0, sticky="we", **pad)
        frm_plate.grid_columnconfigure(0, weight=1)
        frm_plate.grid_columnconfigure(1, weight=1)

        self.plate_var = tk.StringVar()
        self.plate_entry = tb.Entry(frm_plate, textvariable=self.plate_var)
        self.plate_list = tb.Combobox(frm_plate, values=[], state="readonly")
        self.trailer_list = tb.Combobox(frm_plate, values=[], state="readonly")

        tb.Label(frm_plate, text="Номер (для частного лица — вручную):").grid(row=0, column=0, sticky=NW, padx=4,
                                                                              pady=2)
        self.plate_entry.grid(row=1, column=0, sticky="we", padx=4, pady=2)
        self.plate_list.grid(row=1, column=1, sticky="we", padx=4, pady=2)
        tb.Label(frm_plate, text="Номер прицепа (опционально):").grid(row=2, column=0, columnspan=2, sticky=NW, padx=4,
                                                                      pady=2)
        self.trailer_list.grid(row=3, column=0, columnspan=2, sticky="we", padx=4, pady=2)

        # Водитель
        frm_driver = tb.Labelframe(page1, text="Ф.И.О. водителя", padding=10)
        frm_driver.grid(row=2, column=0, sticky="we", **pad)
        self.driver_name = tk.StringVar()
        e = tb.Entry(frm_driver, textvariable=self.driver_name)
        e.grid(row=0, column=0, sticky="we", padx=4, pady=2)
        frm_driver.grid_columnconfigure(0, weight=1)

        # Дефект
        frm_defect = tb.Labelframe(page1, text="Описание заказа и дефекта", padding=10)
        frm_defect.grid(row=3, column=0, sticky="we", **pad)
        frm_defect.grid_columnconfigure(1, weight=1)
        self.defect_choice = tk.StringVar(value=DEFECTS[0])  # "Пропустить" будет по умолчанию
        tb.Label(frm_defect, text="Из списка:").grid(row=0, column=0, sticky=NW, padx=4, pady=2)
        cmb_def = tb.Combobox(frm_defect, textvariable=self.defect_choice, values=DEFECTS, state="readonly")
        cmb_def.grid(row=0, column=1, sticky="we", padx=4, pady=2)
        tb.Label(frm_defect, text="Или 'Другое':").grid(row=1, column=0, sticky=NW, padx=4, pady=2)
        self.defect_custom = tk.StringVar()
        self.defect_entry = tb.Entry(frm_defect, textvariable=self.defect_custom, state=DISABLED)
        self.defect_entry.grid(row=1, column=1, sticky="we", padx=4, pady=2)

        def on_defect_changed(*_):
            if self.defect_choice.get() == "Другое (ввести вручную)":
                self.defect_entry.configure(state=NORMAL)
                self.defect_entry.focus_set()
            else:
                self.defect_entry.configure(state=DISABLED)
                self.defect_custom.set("")

        cmb_def.bind("<<ComboboxSelected>>", lambda e: on_defect_changed())
        on_defect_changed()

        # Исполнители
        frm_people = tb.Labelframe(page1, text="Исполнители", padding=10)
        frm_people.grid(row=4, column=0, sticky="we", **pad)
        frm_people.grid_columnconfigure(1, weight=1)
        self.issued_to = tk.StringVar()
        tb.Label(frm_people, text="Наряд выдан (фамилия исполнителя):").grid(row=0, column=0, sticky=NW, padx=4, pady=2)
        tb.Entry(frm_people, textvariable=self.issued_to).grid(row=0, column=1, sticky="we", padx=4, pady=2)

    def _init_page2(self):
        """Инициализация страницы 2: Услуги и формирование"""
        page2 = self.page2
        page2.grid_columnconfigure(0, weight=1)  # Левая колонка
        page2.grid_columnconfigure(1, weight=1)  # Правая колонка
        page2.grid_rowconfigure(0, weight=1)

        pad = {'padx': 8, 'pady': 6}

        # ===== ЛЕВАЯ КОЛОНКА =====
        left_frame = tb.Frame(page2)
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 4), pady=4)
        left_frame.grid_columnconfigure(0, weight=1)

        # Тип автомобиля
        frm_vehicle = tb.Labelframe(left_frame, text="Тип автомобиля", padding=12)
        frm_vehicle.grid(row=0, column=0, sticky="we", **pad)
        self.vehicle_type = tk.StringVar(value="Легковой")
        tb.Radiobutton(frm_vehicle, text="Легковой", variable=self.vehicle_type,
                       value="Легковой", command=self._on_vehicle_type_changed).pack(side=LEFT, padx=4)
        tb.Radiobutton(frm_vehicle, text="Грузовой", variable=self.vehicle_type,
                       value="Грузовой", command=self._on_vehicle_type_changed).pack(side=LEFT, padx=4)
        tb.Radiobutton(frm_vehicle, text="Спецтехника", variable=self.vehicle_type,
                       value="Спецтехника", command=self._on_vehicle_type_changed).pack(side=LEFT, padx=4)

        # Выбор типа колес
        frm_wheel = tb.Labelframe(left_frame, text="Тип колес", padding=12)
        frm_wheel.grid(row=1, column=0, sticky="we", **pad)
        frm_wheel.grid_columnconfigure(0, weight=1)

        tb.Label(frm_wheel, text="Выберите тип колес:").grid(row=0, column=0, sticky=NW, padx=4, pady=4)
        self.wheel_type = tk.StringVar(value="")
        self.wheel_type_combo = tb.Combobox(frm_wheel, textvariable=self.wheel_type, state="readonly")
        self.wheel_type_combo.grid(row=1, column=0, sticky="we", padx=4, pady=4)

        # Привязка события изменения выбора типа колес
        def on_wheel_type_changed(*_):
            self._update_service_prices()

        self.wheel_type.trace_add("write", on_wheel_type_changed)
        self.wheel_type_combo.bind("<<ComboboxSelected>>", lambda e: self._update_service_prices())

        # === Блок для выбора колес ===
        frm_wheels = tb.Labelframe(left_frame, text="Шины и Диски", padding=12)
        frm_wheels.grid(row=2, column=0, sticky="nsew", **pad)
        frm_wheels.grid_columnconfigure(0, weight=1)
        frm_wheels.grid_columnconfigure(1, weight=1)
        frm_wheels.grid_rowconfigure(3, weight=1)  # Для списка колес

        tb.Label(frm_wheels, text="Выберите колесо:").grid(row=0, column=0, sticky=NW, padx=4, pady=4)
        self.wheel_selected = tk.StringVar(value=WHEELS_LIST[0] if WHEELS_LIST else "")
        self.wheel_combo = tb.Combobox(frm_wheels, textvariable=self.wheel_selected, values=WHEELS_LIST,
                                       state="readonly")
        self.wheel_combo.grid(row=0, column=1, sticky="we", padx=4, pady=4)

        tb.Label(frm_wheels, text="Количество:").grid(row=1, column=0, sticky=NW, padx=4, pady=4)
        self.wheel_quantity = tk.IntVar(value=1)
        tb.Spinbox(frm_wheels, from_=1, to=999, textvariable=self.wheel_quantity, width=10).grid(row=1, column=1,
                                                                                                 sticky="w", padx=4,
                                                                                                 pady=4)

        tb.Button(frm_wheels, text="Добавить колесо", bootstyle="success",
                  command=self._add_wheel).grid(row=2, column=0, columnspan=2, sticky="we", padx=4, pady=4)

        # Список добавленных колес
        self.added_wheels_listbox = tk.Listbox(frm_wheels, height=8)
        self.added_wheels_listbox.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=4, pady=4)

        tb.Button(frm_wheels, text="Удалить выбранное", bootstyle="danger",
                  command=self._remove_wheel).grid(row=4, column=0, columnspan=2, sticky="we", padx=4, pady=4)

        # ===== ПРАВАЯ КОЛОНКА =====
        right_frame = tb.Frame(page2)
        right_frame.grid(row=0, column=1, sticky="nsew", padx=(4, 0), pady=4)
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(0, weight=1)

        # Услуги
        frm_services = tb.Labelframe(right_frame, text="Услуги", padding=12)
        frm_services.grid(row=0, column=0, sticky="nsew", **pad)
        frm_services.grid_columnconfigure(0, weight=1)
        frm_services.grid_rowconfigure(1, weight=1)

        # Шапка услуг
        header = tb.Frame(frm_services)
        header.grid(row=0, column=0, sticky="we")
        header.grid_columnconfigure(0, weight=1)
        tb.Label(header, text="Услуга").grid(row=0, column=0, sticky="w", padx=4, pady=2)
        tb.Label(header, text="Кол-во").grid(row=0, column=1, sticky="w", padx=4, pady=2)
        tb.Label(header, text="Цена (шт)").grid(row=0, column=2, sticky="w", padx=4, pady=2)

        # Прокручиваемый список услуг
        svc_scroll = VScrollFrame(frm_services)
        svc_scroll.grid(row=1, column=0, sticky="nsew", pady=(4, 0))
        svc_inner = svc_scroll.inner

        self.services_vars = {}
        self.services_qty = {}
        self.service_price_labels = {}
        for i, name in enumerate(SERVICES, start=1):
            var = tk.IntVar(value=0)
            qty = tk.IntVar(value=0)

            def _on_toggle_factory(v=var, q=qty):
                def handler():
                    if v.get() and q.get() == 0:
                        q.set(1)
                    if not v.get():
                        q.set(0)

                return handler

            row_frame = tb.Frame(svc_inner)
            row_frame.grid(row=i, column=0, sticky="we", pady=2)
            row_frame.grid_columnconfigure(0, weight=1)

            tb.Checkbutton(row_frame, text=name, variable=var, command=_on_toggle_factory()).grid(
                row=0, column=0, sticky="w", padx=4, ipadx=4)
            tb.Spinbox(row_frame, from_=0, to=999, textvariable=qty, width=8).grid(
                row=0, column=1, sticky="w", padx=4)
            lbl = tb.Label(row_frame, text="-", width=10, anchor="w")
            lbl.grid(row=0, column=2, sticky="w", padx=4)

            self.services_vars[name] = var
            self.services_qty[name] = qty
            self.service_price_labels[name] = lbl

        # Инициализация списка для хранения добавленных колес
        self.added_wheels = []

        # Инициализация типов колес
        self._on_vehicle_type_changed()
        self._update_service_prices()

    def _go_to_page1(self):
        """Переход на страницу 1"""
        self.page2.grid_forget()
        self.page1.grid(row=0, column=0, sticky="nsew")
        self.current_page = 0
        self.btn_back.configure(state=DISABLED)
        self.btn_next.configure(state=NORMAL)
        self.btn_create.pack_forget()
        self.btn_next.pack(side=RIGHT, padx=4)
        self.page_label.configure(text="Страница 1/2")

    def _go_to_page2(self):
        """Переход на страницу 2 с проверкой данных страницы 1"""
        # Проверяем обязательные поля страницы 1
        if not self._validate_page1():
            return

        self.page1.grid_forget()
        self.page2.grid(row=0, column=0, sticky="nsew")
        self.current_page = 1
        self.btn_back.configure(state=NORMAL)
        self.btn_next.pack_forget()
        self.btn_create.pack(side=RIGHT, padx=4)
        self.page_label.configure(text="Страница 2/2")

        # Обновляем цены услуг при переходе
        self._update_service_prices()

    def _validate_page1(self):
        """Проверка обязательных полей на странице 1"""
        if self.customer_type.get() == "Компания":
            if not self.company_selected.get():
                messagebox.showerror("Ошибка", "Выберите компанию.", parent=self._form_parent)
                return False
            if self.company_selected.get() not in ALL_COMPANY_NAMES:
                messagebox.showerror("Ошибка", "Компания недоступна (возможно, Оплата=нет).", parent=self._form_parent)
                return False
            if not self.plate_list.get():
                messagebox.showerror("Ошибка", "Выберите гос. номер из списка.", parent=self._form_parent)
                return False
        else:
            if not self.plate_entry.get().strip():
                messagebox.showerror("Ошибка", "Введите гос. номер для частного лица.", parent=self._form_parent)
                return False

        if not self.driver_name.get().strip():
            messagebox.showerror("Ошибка", "Введите Ф.И.О. водителя.", parent=self._form_parent)
            return False

        if self.defect_choice.get() == "Другое (ввести вручную)":
            if not self.defect_custom.get().strip():
                messagebox.showerror("Ошибка", "Введите текст дефекта в поле 'Другое'.", parent=self._form_parent)
                return False

        if not self.issued_to.get().strip():
            messagebox.showerror("Ошибка", "Введите фамилию исполнителя ('Наряд выдан').", parent=self._form_parent)
            return False

        return True

    def _add_wheel(self):
        """Добавляет выбранное колесо в список"""
        wheel_name = self.wheel_selected.get()
        quantity = self.wheel_quantity.get()

        if not wheel_name:
            messagebox.showwarning("Предупреждение", "Выберите колесо из списка.", parent=self._form_parent)
            return

        if quantity <= 0:
            messagebox.showwarning("Предупреждение", "Укажите количество больше 0.", parent=self._form_parent)
            return

        # Формируем запись в формате "2x - Название колеса"
        wheel_entry = f"{quantity}x - {wheel_name}"
        self.added_wheels.append(wheel_entry)
        self.added_wheels_listbox.insert(tk.END, wheel_entry)

        # Очищаем выбор и сбрасываем количество
        self.wheel_quantity.set(1)

    def _remove_wheel(self):
        """Удаляет выбранное колесо из списка"""
        selection = self.added_wheels_listbox.curselection()
        if not selection:
            messagebox.showwarning("Предупреждение", "Выберите колесо для удаления.", parent=self._form_parent)
            return

        index = selection[0]
        self.added_wheels_listbox.delete(index)
        self.added_wheels.pop(index)

    # Применить текущий справочник к открытой форме
    def _apply_companies_to_form(self, win):
        # форма может быть не открытой или уже закрыта
        if not hasattr(self, "cmb_company") or not self._widget_exists(self.cmb_company):
            return
        self.cmb_company["values"] = ALL_COMPANY_NAMES
        if ALL_COMPANY_NAMES:
            self.cmb_company.set(ALL_COMPANY_NAMES[0])
        else:
            self.cmb_company.set("")
        # перезаполнить поиск (если виджеты живы)
        if hasattr(self, "company_query"):
            q = self.company_query.get()
            values = filter_companies(q)
            self.cmb_company["values"] = values
            if values:
                self.cmb_company.set(values[0])
            else:
                self.cmb_company.set("")
            if hasattr(self, "search_results") and self._widget_exists(self.search_results):
                self.search_results.set_items(values[:50], q.strip().lower())
        self._update_company_meta()

    def _on_vehicle_type_changed(self):
        """Обновляет список доступных типов колес при смене типа автомобиля"""
        vehicle_type = self.vehicle_type.get()
        wheel_types = WHEEL_TYPES.get(vehicle_type, [])

        self.wheel_type_combo["values"] = wheel_types
        if wheel_types:
            self.wheel_type.set(wheel_types[0])
        else:
            self.wheel_type.set("")

        self._update_service_prices()  # Обновляем цены при смене типа авто

    # ======= Админ‑панель =======
    def open_admin_panel(self):
        # пароль
        pwd = simpledialog.askstring("Вход в админ‑панель", "Введите пароль:", show='*', parent=self.root)
        if pwd != "12345":
            messagebox.showerror("Доступ запрещён", "Неверный пароль.", parent=self.root)
            return

        win = tb.Toplevel(self.root)
        win.title("Админ‑панель")
        win.geometry("1000x700")
        nb = ttk.Notebook(win)
        nb.pack(fill=BOTH, expand=True, padx=8, pady=8)

        # ====== вкладка Добавить компанию ======
        tab_add_company = tb.Frame(nb, padding=10)
        nb.add(tab_add_company, text="Добавить компанию")

        name_var = tk.StringVar()
        inn_var = tk.StringVar()
        plates_var = tk.StringVar()
        tb.Label(tab_add_company, text="Название компании:").grid(row=0, column=0, sticky=NW, pady=4)
        tb.Entry(tab_add_company, textvariable=name_var).grid(row=0, column=1, sticky="we", pady=4)
        tb.Label(tab_add_company, text="ИНН:").grid(row=1, column=0, sticky=NW, pady=4)
        tb.Entry(tab_add_company, textvariable=inn_var).grid(row=1, column=1, sticky="we", pady=4)
        tb.Label(tab_add_company, text="Гос. номера (через запятую):").grid(row=2, column=0, sticky=NW, pady=4)
        tb.Entry(tab_add_company, textvariable=plates_var).grid(row=2, column=1, sticky="we", pady=4)
        tab_add_company.grid_columnconfigure(1, weight=1)

        def do_add_company():
            name = name_var.get().strip()
            inn = inn_var.get().strip()
            plates = join_plates(parse_plates(plates_var.get()))
            if not name:
                messagebox.showerror("Ошибка", "Введите название компании.", parent=win);
                return
            df = read_companies_df()
            if (df[COL_NAME].str.lower() == name.lower()).any():
                messagebox.showerror("Ошибка", "Компания с таким названием уже существует.", parent=win);
                return
            # добавляем В КОНЕЦ с НДС по умолчанию "нет"
            df.loc[len(df)] = {COL_NAME: name, COL_INN: inn, COL_PLATES: plates, COL_PAY: "да", COL_VAT: "нет"}
            write_companies_df(df)
            reload_companies_globals()
            # обновим GUI, если окно формы открыто
            self._apply_companies_to_form(self._create_form_window)
            # обновим списки во всех вкладках админки
            _apply_filter1();
            _apply_filter2();
            _apply_filter3();
            _apply_filter4();
            _refresh_plates_list();
            _sync_pay_toggle();
            _sync_vat_toggle()
            messagebox.showinfo("Готово", "Компания добавлена (в конец) и включена в списки (Оплата=да, НДС=нет).",
                                parent=win)

        tb.Button(tab_add_company, text="Добавить", bootstyle="success", command=do_add_company).grid(row=3, column=1,
                                                                                                      sticky="e",
                                                                                                      pady=8)

        # ====== вкладка Добавить гос.номер ======
        tab_add_plate = tb.Frame(nb, padding=10)
        nb.add(tab_add_plate, text="Добавить гос.номер")

        q1 = tk.StringVar()
        tb.Label(tab_add_plate, text="Поиск компании:").grid(row=0, column=0, sticky=NW, pady=4)
        e_q1 = tb.Entry(tab_add_plate, textvariable=q1);
        e_q1.grid(row=0, column=1, sticky="we", pady=4)
        tab_add_plate.grid_columnconfigure(1, weight=1)
        combo1 = tb.Combobox(tab_add_plate, values=list(COMPANIES.keys()), state="readonly")
        combo1.grid(row=1, column=0, columnspan=2, sticky="we", pady=4)

        def _apply_filter1(*_):
            all_names = list(COMPANIES.keys())
            qq = q1.get().strip().lower()
            vals = [n for n in all_names if qq in n.lower()]
            combo1["values"] = vals
            if vals:
                combo1.set(vals[0])

        q1.trace_add("write", _apply_filter1)
        _apply_filter1()

        newplates_var = tk.StringVar()
        tb.Label(tab_add_plate, text="Новые номера (через запятую):").grid(row=2, column=0, sticky=NW, pady=4)
        tb.Entry(tab_add_plate, textvariable=newplates_var).grid(row=2, column=1, sticky="we", pady=4)

        def do_add_plates():
            name = combo1.get().strip()
            if not name:
                messagebox.showerror("Ошибка", "Выберите компанию.", parent=win);
                return
            df = read_companies_df()
            mask = df[COL_NAME].str.lower() == name.lower()
            if not mask.any():
                messagebox.showerror("Ошибка", "Компания не найдена в таблице.", parent=win);
                return
            plates_old = parse_plates(df.loc[mask, COL_PLATES].iloc[0])
            plates_new = parse_plates(newplates_var.get())
            plates_joined = join_plates(plates_old + plates_new)
            df.loc[mask, COL_PLATES] = plates_joined
            write_companies_df(df)
            reload_companies_globals()
            self._apply_companies_to_form(self._create_form_window)
            _apply_filter1();
            _refresh_plates_list()
            messagebox.showinfo("Готово", "Номера добавлены.", parent=win)

        tb.Button(tab_add_plate, text="Добавить номера", bootstyle="success", command=do_add_plates).grid(row=3,
                                                                                                          column=1,
                                                                                                          sticky="e",
                                                                                                          pady=8)

        # ====== вкладка Оплата on/off ======
        tab_pay = tb.Frame(nb, padding=10)
        nb.add(tab_pay, text="Выставить оплату")

        q2 = tk.StringVar()
        tb.Label(tab_pay, text="Поиск компании:").grid(row=0, column=0, sticky=NW, pady=4)
        e_q2 = tb.Entry(tab_pay, textvariable=q2);
        e_q2.grid(row=0, column=1, sticky="we", pady=4)
        tab_pay.grid_columnconfigure(1, weight=1)
        combo2 = tb.Combobox(tab_pay, values=list(COMPANIES.keys()), state="readonly")
        combo2.grid(row=1, column=0, columnspan=2, sticky="we", pady=4)

        pay_var = tk.BooleanVar(value=False)
        tb.Checkbutton(tab_pay, text="Оплата включена (да)", variable=pay_var, bootstyle="success-square-toggle").grid(
            row=2, column=0, sticky=NW, pady=4)

        def _sync_pay_toggle(*_):
            name = combo2.get().strip()
            if not name:
                pay_var.set(False);
                return
            df_state = read_companies_df()
            mask = df_state[COL_NAME].str.lower() == name.lower()
            current = str(df_state.loc[mask, COL_PAY].iloc[0]).strip().lower() if mask.any() else ''
            pay_var.set(current in ("да", "yes", "true", "1"))

        def _apply_filter2(*_):
            all_names = list(COMPANIES.keys())
            qq = q2.get().strip().lower()
            vals = [n for n in all_names if qq in n.lower()]
            combo2["values"] = vals
            if vals:
                combo2.set(vals[0])
                _sync_pay_toggle()

        q2.trace_add("write", _apply_filter2);
        _apply_filter2()
        combo2.bind("<<ComboboxSelected>>", _sync_pay_toggle)

        def do_set_pay():
            name = combo2.get().strip()
            df = read_companies_df()
            mask = df[COL_NAME].str.lower() == name.lower()
            if not mask.any():
                messagebox.showerror("Ошибка", "Компания не найдена.", parent=win);
                return
            df.loc[mask, COL_PAY] = "да" if pay_var.get() else "нет"
            write_companies_df(df)
            reload_companies_globals()
            self._apply_companies_to_form(self._create_form_window)
            _apply_filter2();
            _sync_pay_toggle()
            messagebox.showinfo("Готово", "Статус оплаты обновлён.", parent=win)

        tb.Button(tab_pay, text="Сохранить", bootstyle="success", command=do_set_pay).grid(row=3, column=1, sticky="e",
                                                                                           pady=8)

        # ====== вкладка НДС on/off ======
        tab_vat = tb.Frame(nb, padding=10)
        nb.add(tab_vat, text="Выставить НДС")

        q_vat = tk.StringVar()
        tb.Label(tab_vat, text="Поиск компании:").grid(row=0, column=0, sticky=NW, pady=4)
        e_q_vat = tb.Entry(tab_vat, textvariable=q_vat);
        e_q_vat.grid(row=0, column=1, sticky="we", pady=4)
        tab_vat.grid_columnconfigure(1, weight=1)
        combo_vat = tb.Combobox(tab_vat, values=list(COMPANIES.keys()), state="readonly")
        combo_vat.grid(row=1, column=0, columnspan=2, sticky="we", pady=4)

        vat_var = tk.BooleanVar(value=False)
        tb.Checkbutton(tab_vat, text="НДС включен (да)", variable=vat_var, bootstyle="success-square-toggle").grid(
            row=2, column=0, sticky=NW, pady=4)

        def _sync_vat_toggle(*_):
            name = combo_vat.get().strip()
            if not name:
                vat_var.set(False);
                return
            df_state = read_companies_df()
            mask = df_state[COL_NAME].str.lower() == name.lower()
            current = str(df_state.loc[mask, COL_VAT].iloc[0]).strip().lower() if mask.any() else ''
            vat_var.set(current in ("да", "yes", "true", "1"))

        def _apply_filter_vat(*_):
            all_names = list(COMPANIES.keys())
            qq = q_vat.get().strip().lower()
            vals = [n for n in all_names if qq in n.lower()]
            combo_vat["values"] = vals
            if vals:
                combo_vat.set(vals[0])
                _sync_vat_toggle()

        q_vat.trace_add("write", _apply_filter_vat);
        _apply_filter_vat()
        combo_vat.bind("<<ComboboxSelected>>", _sync_vat_toggle)

        def do_set_vat():
            name = combo_vat.get().strip()
            df = read_companies_df()
            mask = df[COL_NAME].str.lower() == name.lower()
            if not mask.any():
                messagebox.showerror("Ошибка", "Компания не найдена.", parent=win);
                return
            df.loc[mask, COL_VAT] = "да" if vat_var.get() else "нет"
            write_companies_df(df)
            reload_companies_globals()
            self._apply_companies_to_form(self._create_form_window)
            _apply_filter_vat();
            _sync_vat_toggle()
            messagebox.showinfo("Готово", "Статус НДС обновлён.", parent=win)

        tb.Button(tab_vat, text="Сохранить", bootstyle="success", command=do_set_vat).grid(row=3, column=1, sticky="e",
                                                                                           pady=8)

        # ====== вкладка Удалить компанию ======
        tab_del_company = tb.Frame(nb, padding=10)
        nb.add(tab_del_company, text="Удалить компанию")

        q3 = tk.StringVar()
        tb.Label(tab_del_company, text="Поиск компании:").grid(row=0, column=0, sticky=NW, pady=4)
        e_q3 = tb.Entry(tab_del_company, textvariable=q3);
        e_q3.grid(row=0, column=1, sticky="we", pady=4)
        tab_del_company.grid_columnconfigure(1, weight=1)
        combo3 = tb.Combobox(tab_del_company, values=list(COMPANIES.keys()), state="readonly")
        combo3.grid(row=1, column=0, columnspan=2, sticky="we", pady=4)

        def _apply_filter3(*_):
            all_names = list(COMPANIES.keys())
            qq = q3.get().strip().lower()
            vals = [n for n in all_names if qq in n.lower()]
            combo3["values"] = vals
            if vals:
                combo3.set(vals[0])

        q3.trace_add("write", _apply_filter3);
        _apply_filter3()

        def do_del_company():
            name = combo3.get().strip()
            if not name:
                messagebox.showerror("Ошибка", "Выберите компанию.", parent=win);
                return
            if not messagebox.askyesno("Подтвердите", f"Удалить компанию «{name}» и все её номера?", parent=win):
                return
            df = read_companies_df()
            df = df[~(df[COL_NAME].str.lower() == name.lower())]
            write_companies_df(df)
            reload_companies_globals()
            self._apply_companies_to_form(self._create_form_window)
            _apply_filter1();
            _apply_filter2();
            _apply_filter3();
            _apply_filter4();
            _refresh_plates_list();
            _sync_pay_toggle();
            _sync_vat_toggle()
            messagebox.showinfo("Готово", "Компания удалена.", parent=win)

        tb.Button(tab_del_company, text="Удалить", bootstyle="danger", command=do_del_company).grid(row=2, column=1,
                                                                                                    sticky="e", pady=8)

        # ====== вкладка Удалить гос.номер ======
        tab_del_plate = tb.Frame(nb, padding=10)
        nb.add(tab_del_plate, text="Удалить гос. номер")

        q4 = tk.StringVar()
        tb.Label(tab_del_plate, text="Поиск компании:").grid(row=0, column=0, sticky=NW, pady=4)
        e_q4 = tb.Entry(tab_del_plate, textvariable=q4);
        e_q4.grid(row=0, column=1, sticky="we", pady=4)
        tab_del_plate.grid_columnconfigure(1, weight=1)
        combo4 = tb.Combobox(tab_del_plate, values=list(COMPANIES.keys()), state="readonly")
        combo4.grid(row=1, column=0, columnspan=2, sticky="we", pady=4)

        listbox = tk.Listbox(tab_del_plate, selectmode="extended", height=12)
        listbox.grid(row=2, column=0, columnspan=2, sticky="nsew", pady=6)
        tab_del_plate.grid_rowconfigure(2, weight=1)

        def _refresh_plates_list(*_):
            name = combo4.get().strip()
            listbox.delete(0, tk.END)
            if name and name in COMPANIES:
                for p in COMPANIES[name]["plates"]:
                    listbox.insert(tk.END, p)

        def _apply_filter4(*_):
            all_names = list(COMPANIES.keys())
            qq = q4.get().strip().lower()
            vals = [n for n in all_names if qq in n.lower()]
            combo4["values"] = vals
            if vals:
                combo4.set(vals[0])
                _refresh_plates_list()

        q4.trace_add("write", _apply_filter4);
        _apply_filter4()
        combo4.bind("<<ComboboxSelected>>", lambda e: _refresh_plates_list())

        def do_del_plates():
            name = combo4.get().strip()
            if not name:
                messagebox.showerror("Ошибка", "Выберите компанию.", parent=win);
                return
            sel = [listbox.get(i) for i in listbox.curselection()]
            if not sel:
                messagebox.showerror("Ошибка", "Выберите номера для удаления.", parent=win);
                return
            df = read_companies_df()
            mask = df[COL_NAME].str.lower() == name.lower()
            if not mask.any():
                messagebox.showerror("Ошибка", "Компания не найдена в таблице.", parent=win);
                return
            old = parse_plates(df.loc[mask, COL_PLATES].iloc[0])
            new = [p for p in old if p not in sel]
            df.loc[mask, COL_PLATES] = join_plates(new)
            write_companies_df(df)
            reload_companies_globals()
            self._apply_companies_to_form(self._create_form_window)
            _apply_filter4();
            _refresh_plates_list()
            messagebox.showinfo("Готово", "Выбранные номера удалены.", parent=win)

        tb.Button(tab_del_plate, text="Удалить отмеченные номера", bootstyle="danger", command=do_del_plates).grid(
            row=3, column=1, sticky="e", pady=8)

    # ======= ЛОГИКА формы =======
    def _widget_exists(self, w) -> bool:
        try:
            return bool(w and w.winfo_exists())
        except Exception:
            return False

    def _update_company_meta(self):
        name = getattr(self, "company_selected", tk.StringVar()).get()
        meta = COMPANIES.get(name, {"inn": "", "cars": [], "trailers": [], "plates": [], "vat": "нет"})
        if hasattr(self, "company_inn_var"):
            self.company_inn_var.set(meta.get("inn", ""))
        if hasattr(self, "company_vat_var"):
            vat_status = meta.get("vat", "нет")
            self.company_vat_var.set(f"НДС: {'да' if vat_status in ('да', 'yes', 'true', '1') else 'нет'}")
        q = ""
        if hasattr(self, "company_query"):
            q = self.company_query.get().strip().lower()
        if hasattr(self, "plate_list") and self._widget_exists(self.plate_list):
            cars = meta.get("cars", [])
            self.plate_list["values"] = cars
            sel_plate = ""
            for p in cars:
                if q and q in p.lower():
                    sel_plate = p
                    break
            if sel_plate:
                self.plate_list.set(sel_plate)
            elif cars:
                self.plate_list.set(cars[0])
            else:
                self.plate_list.set("")
        if hasattr(self, "trailer_list") and self._widget_exists(self.trailer_list):
            trailers = ["Без прицепа"] + meta.get("trailers", [])
            self.trailer_list["values"] = trailers
            sel_trailer = ""
            for t in trailers:
                if q and q in t.lower():
                    sel_trailer = t
                    break
            if sel_trailer:
                self.trailer_list.set(sel_trailer)
            elif trailers:
                self.trailer_list.set(trailers[0])
            else:
                self.trailer_list.set("")

    def _on_customer_type_changed(self):
        is_company = (self.customer_type.get() == "Компания")
        if hasattr(self, "plate_entry") and hasattr(self, "plate_list"):
            if is_company:
                self.plate_entry.configure(state=DISABLED)
                self.plate_list.configure(state="readonly")
                if hasattr(self, "trailer_list"):
                    self.trailer_list.configure(state="readonly")
            else:
                self.plate_entry.configure(state=NORMAL)
                self.plate_list.configure(state=DISABLED)
                if hasattr(self, "trailer_list"):
                    self.trailer_list.configure(state=DISABLED)

    def _update_service_prices(self):
        """Обновляет цены услуг с учетом типа автомобиля и типа колес"""
        vt = self.vehicle_type.get()
        wt = self.wheel_type.get()

        # Определяем, использовать ли цены с НДС
        use_nds = False
        if self.customer_type.get() == "Компания":
            company_name = getattr(self, "company_selected", tk.StringVar()).get()
            if company_name:
                meta = COMPANIES.get(company_name, {})
                vat_status = meta.get("vat", "нет")
                use_nds = (vat_status in ("да", "yes", "true", "1"))

        # Выбираем правильную таблицу цен
        price_table = PRICE_TABLE_NDS if use_nds else PRICE_TABLE

        if not price_table.get(vt):
            price_table = load_price_table(use_nds=use_nds)

        for name, lbl in getattr(self, "service_price_labels", {}).items():
            base_name = SERVICE_PRICE_NAME.get(name, name)
            price_key = f"{base_name}|{wt}" if wt else base_name

            # Ищем цену с учетом типа колес
            price = price_table.get(vt, {}).get(price_key)

            # Если не нашли с типом колес, ищем без типа
            if price is None:
                price = price_table.get(vt, {}).get(base_name, "-")

            if isinstance(price, tuple):
                lbl.configure(text=f"{price[0]}/{price[1]}")
            elif price and price != 0:
                lbl.configure(text=str(price))
            else:
                lbl.configure(text="-")

    def _ask_split_service(self, title: str, labels: list[str], total: int) -> list[int]:
        win = tb.Toplevel(self._form_parent)
        win.title(title)
        vars = []
        for i, lab in enumerate(labels):
            row = tb.Frame(win, padding=4)
            row.grid(row=i, column=0)
            tb.Label(row, text=lab).pack(side=LEFT, padx=4)
            val = tk.IntVar(value=(total if i == 0 else 0))
            tb.Spinbox(row, from_=0, to=999, textvariable=val, width=6).pack(side=LEFT, padx=4)
            vars.append(val)
        res = []

        def _ok():
            for v in vars:
                res.append(int(v.get()))
            win.destroy()

        tb.Button(win, text="OK", command=_ok).grid(row=len(labels), column=0, pady=6)
        self._form_parent.wait_window(win)
        return res

    # В класс WorkOrderApp добавляем новый метод для диалога грузиков
    def _ask_weights_dialog(self, kind: str, total_qty: int, use_nds: bool = False):
        """Диалог для выбора грузиков с возможностью выбрать несколько типов"""
        win = tb.Toplevel(self._form_parent)
        win.title(f"Выбор грузиков ({total_qty} шт.)")
        win.geometry("700x800")
        win.grab_set()

        # Используем правильную таблицу расходников
        consumables_table = CONSUMABLES_TABLE_NDS if use_nds else CONSUMABLES_TABLE
        consumable_categories = CONSUMABLE_CATEGORIES_NDS if use_nds else CONSUMABLE_CATEGORIES

        # Получаем доступные названия грузиков
        names = sorted(consumables_table.get(kind, {}).keys())

        # Фрейм для отображения выбранных грузиков
        selected_frame = tb.LabelFrame(win, text="Выбранные грузики", padding=10)
        selected_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Прокручиваемый фрейм для списка выбранных
        selected_scroll = tk.Frame(selected_frame)
        selected_scroll.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(selected_scroll, height=200)
        scrollbar = ttk.Scrollbar(selected_scroll, orient="vertical", command=canvas.yview)
        scrollable_frame = tb.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Список для хранения выбранных грузиков
        selected_items = []  # список словарей: {"name": "", "category": "", "qty": 0}

        def update_selected_list():
            """Обновляет отображение выбранных грузиков"""
            # Очищаем фрейм
            for widget in scrollable_frame.winfo_children():
                widget.destroy()

            # Добавляем выбранные позиции
            for i, item in enumerate(selected_items):
                row_frame = tb.Frame(scrollable_frame)
                row_frame.pack(fill=tk.X, pady=2)

                tb.Label(row_frame, text=f"{i + 1}. {item['name']} ({item['category']}) - {item['qty']} шт.",
                         width=50, anchor="w").pack(side=tk.LEFT, padx=5)

                # Кнопка удалить
                tb.Button(row_frame, text="✕", bootstyle="danger-outline", width=3,
                          command=lambda idx=i: remove_selected(idx)).pack(side=tk.RIGHT)

            # Обновляем общее количество
            total_selected = sum(item['qty'] for item in selected_items)
            remaining = total_qty - total_selected
            status_label.config(text=f"Осталось выбрать: {remaining} шт. | Всего выбрано: {total_selected} шт.")

            # Активируем/деактивируем кнопку ОК
            if remaining == 0:
                ok_button.configure(state=tk.NORMAL)
            else:
                ok_button.configure(state=tk.DISABLED)

        def remove_selected(idx):
            """Удаляет выбранный грузик"""
            selected_items.pop(idx)
            update_selected_list()

        # Фрейм для добавления нового грузика
        add_frame = tb.LabelFrame(win, text="Добавить грузик", padding=10)
        add_frame.pack(fill=tk.X, padx=10, pady=5)

        # Название
        tb.Label(add_frame, text="Название:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        name_var = tk.StringVar(value=names[0] if names else "")
        name_combo = tb.Combobox(add_frame, textvariable=name_var, values=names,
                                 state="readonly", width=30)
        name_combo.grid(row=0, column=1, padx=5, pady=5)

        # Категория
        tb.Label(add_frame, text="Категория:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        category_var = tk.StringVar(value=consumable_categories[0] if consumable_categories else "")
        category_combo = tb.Combobox(add_frame, textvariable=category_var,
                                     values=consumable_categories, state="readonly", width=30)
        category_combo.grid(row=1, column=1, padx=5, pady=5)

        # Количество
        tb.Label(add_frame, text="Количество:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        qty_var = tk.IntVar(value=1)
        qty_spin = tb.Spinbox(add_frame, from_=1, to=total_qty, textvariable=qty_var, width=10)
        qty_spin.grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)

        def add_weight():
            """Добавляет выбранный грузик в список"""
            name = name_var.get()
            category = category_var.get()
            qty = qty_var.get()

            if not name or not category:
                messagebox.showwarning("Ошибка", "Выберите название и категорию", parent=win)
                return

            # Проверяем общее количество
            total_selected = sum(item['qty'] for item in selected_items)
            if total_selected + qty > total_qty:
                messagebox.showwarning("Ошибка",
                                       f"Нельзя выбрать больше {total_qty} шт. Всего уже выбрано: {total_selected} шт.",
                                       parent=win)
                return

            # Добавляем в список
            selected_items.append({
                "name": name,
                "category": category,
                "qty": qty,
                "temperature": "холодная"  # Для грузиков всегда холодная
            })

            update_selected_list()

            # Сбрасываем количество на минимальное
            remaining = total_qty - sum(item['qty'] for item in selected_items)
            qty_var.set(min(1, remaining))

        # Кнопка добавить
        add_button = tb.Button(add_frame, text="Добавить грузик", bootstyle="success",
                               command=add_weight)
        add_button.grid(row=3, column=1, sticky=tk.E, padx=5, pady=10)

        # Статус
        status_frame = tb.Frame(win)
        status_frame.pack(fill=tk.X, padx=10, pady=5)
        status_label = tb.Label(status_frame, text=f"Нужно выбрать: {total_qty} шт.")
        status_label.pack()

        # Кнопки ОК/Отмена
        button_frame = tb.Frame(win)
        button_frame.pack(fill=tk.X, padx=10, pady=10)

        ok_button = tb.Button(button_frame, text="OK", bootstyle="primary",
                              state=tk.DISABLED, command=win.destroy)
        ok_button.pack(side=tk.RIGHT, padx=5)

        tb.Button(button_frame, text="Отмена", bootstyle="secondary",
                  command=lambda: [selected_items.clear(), win.destroy()]).pack(side=tk.RIGHT, padx=5)

        # Инициализация
        update_selected_list()

        # Ждем закрытия окна
        win.wait_window(win)

        # Возвращаем результат в формате (name, category, temperature, qty)
        result = []
        for item in selected_items:
            # Для каждого выбранного грузика добавляем соответствующее количество записей
            # но с учетом, что в итоге нам нужно общее количество и средняя цена
            result.extend([(item["name"], item["category"], item["temperature"])] * item["qty"])

        return result

    def _ask_consumables(self, kind: str, qty: int):
        # Определяем, использовать ли расходники с НДС
        use_nds = False
        if self.customer_type.get() == "Компания":
            company_name = getattr(self, "company_selected", tk.StringVar()).get()
            if company_name:
                meta = COMPANIES.get(company_name, {})
                vat_status = meta.get("vat", "нет")
                use_nds = (vat_status in ("да", "yes", "true", "1"))

        # перечитываем файл, чтобы гарантировать актуальные данные
        if use_nds:
            global CONSUMABLES_TABLE_NDS, CONSUMABLE_CATEGORIES_NDS
            CONSUMABLES_TABLE_NDS, CONSUMABLE_CATEGORIES_NDS = load_consumables_table(use_nds=True)
        else:
            global CONSUMABLES_TABLE, CONSUMABLE_CATEGORIES
            CONSUMABLES_TABLE, CONSUMABLE_CATEGORIES = load_consumables_table()

        # Для Грибка используем специальный диалог с только названием
        if kind == "Грибок":
            dlg = SimpleConsumableDialog(self._form_parent, kind, qty, use_nds=use_nds)
        # Для грузиков используем новый диалог
        elif kind == "Грузики":
            result = self._ask_weights_dialog(kind, qty, use_nds=use_nds)
            # Возвращаем результат в том же формате, что и раньше
            return result
        else:
            dlg = ConsumableDialog(self._form_parent, kind, qty, use_nds=use_nds)

        self._form_parent.wait_window(dlg)
        return dlg.result or []

    def _collect_services(self) -> dict[str, dict]:
        vt = self.vehicle_type.get()
        wt = self.wheel_type.get()

        # Определяем, использовать ли цены с НДС
        use_nds = False
        if self.customer_type.get() == "Компания":
            company_name = getattr(self, "company_selected", tk.StringVar()).get()
            if company_name:
                meta = COMPANIES.get(company_name, {})
                vat_status = meta.get("vat", "нет")
                use_nds = (vat_status in ("да", "yes", "true", "1"))

        # Выбираем правильную таблицу цен
        price_table = PRICE_TABLE_NDS if use_nds else PRICE_TABLE

        if not price_table.get(vt):
            price_table = load_price_table(use_nds=use_nds)

        selected = {}
        for name in SERVICES:
            var = self.services_vars[name]
            qty = max(0, int(self.services_qty[name].get()))

            if not (var.get() and qty > 0):
                continue
            base_name = SERVICE_PRICE_NAME.get(name, name)
            price_key = f"{base_name}|{wt}" if wt else base_name

            if name == "Снятие/установка":
                outer, inner = self._ask_split_service(name, ["наружное", "внутреннее"], qty)
                price = price_table.get(vt, {}).get(price_key) or price_table.get(vt, {}).get(base_name, (0, 0))
                if isinstance(price, int):
                    price = (price, price)
                cost = outer * price[0] + inner * price[1]
                total_qty = outer + inner
                if total_qty > 0:
                    avg = cost // total_qty
                    selected[name] = {"qty": total_qty, "price": avg, "cost": cost}
                    self.services_qty[name].set(total_qty)
                else:
                    self.services_qty[name].set(0)
            elif name == "Вентиль легковой":
                chrome, black = self._ask_split_service(name, ["хром", "черный"], qty)
                price = price_table.get(vt, {}).get(price_key) or price_table.get(vt, {}).get(base_name, (0, 0))
                cost = chrome * price[0] + black * price[1]
                total_qty = chrome + black
                if total_qty > 0:
                    avg = cost // total_qty
                    selected[name] = {"qty": total_qty, "price": avg, "cost": cost}
                    self.services_qty[name].set(total_qty)
                else:
                    self.services_qty[name].set(0)
            elif name in CONSUMABLE_SERVICE_MAP:
                kind = CONSUMABLE_SERVICE_MAP[name]
                items = self._ask_consumables(kind, qty)

                # ИСПРАВЛЕНО: для каждого расходника считаем отдельно
                if items:
                    # Используем правильную таблицу расходников
                    consumables_table = CONSUMABLES_TABLE_NDS if use_nds else CONSUMABLES_TABLE

                    # Считаем общую стоимость всех выбранных расходников
                    total_cost = 0
                    individual_costs = []  # Для отладки

                    for item in items:
                        if len(item) == 3:
                            n, c, t = item
                        else:
                            # Для Грибка, где возвращается только название
                            n = item[0]
                            # Используем фиксированные значения для Грибка
                            c = "Грузовые автомобили 230-445 мм"
                            t = "холодная"

                        # Ищем цену в таблице расходников
                        item_cost = 0
                        if kind in consumables_table and n in consumables_table[kind]:
                            # Для Грибка ищем конкретную категорию и температуру
                            if kind == "Грибок":
                                # Для грибка ищем конкретную запись
                                price_entry = consumables_table[kind][n].get((c, t))
                                if price_entry:
                                    item_cost = price_entry
                                    individual_costs.append(f"{n}: {price_entry} руб")
                                else:
                                    # Если не нашли, берем первую доступную цену
                                    available_prices = list(consumables_table[kind][n].values())
                                    if available_prices:
                                        item_cost = available_prices[0]
                                        individual_costs.append(f"{n}: {available_prices[0]} руб (первая доступная)")
                            else:
                                # Для остальных расходников используем выбранные значения
                                price_entry = consumables_table[kind][n].get((c, t))
                                if price_entry:
                                    item_cost = price_entry
                                    individual_costs.append(f"{n} ({c}, {t}): {price_entry} руб")

                        total_cost += item_cost

                    # Для отладки
                    if individual_costs:
                        print(f"Расходники '{kind}':")
                        for cost_info in individual_costs:
                            print(f"  {cost_info}")
                        print(f"  Общая стоимость: {total_cost} руб")

                    total_qty = len(items)
                    if total_qty > 0:
                        # ИСПРАВЛЕНО: считаем среднюю цену правильно
                        avg_price = total_cost // total_qty if total_cost > 0 else 0
                        selected[name] = {"qty": total_qty, "price": avg_price, "cost": total_cost}
                        self.services_qty[name].set(total_qty)
            else:
                price = price_table.get(vt, {}).get(price_key) or price_table.get(vt, {}).get(base_name, 0)
                cost = price * qty
                selected[name] = {"qty": qty, "price": price, "cost": cost}
        return selected

    def _validate(self) -> tuple[bool, str]:
        # Проверка обязательных полей страницы 2
        if not self.wheel_type.get():
            return False, "Выберите тип колес."

        if not any(self.services_vars[name].get() and int(self.services_qty[name].get()) > 0 for name in SERVICES):
            return False, "Выберите хотя бы одну услугу и укажите количество."
        return True, ""

    def _gather_data(self) -> dict:
        is_company = (self.customer_type.get() == "Компания")
        if is_company:
            customer_display = self.company_selected.get()
            plate_value = self.plate_list.get().strip()
            trailer_value = self.trailer_list.get().strip() if hasattr(self, "trailer_list") else ""
            if trailer_value == "Без прицепа":
                trailer_value = ""
        else:
            customer_display = "Частное лицо"
            plate_value = self.plate_entry.get().strip()
            trailer_value = ""

        # ИЗМЕНЕНО: если выбрано "Пропустить", оставляем поле пустым
        if self.defect_choice.get() == "Другое (ввести вручную)":
            defect_value = self.defect_custom.get().strip()
        elif self.defect_choice.get() == "Пропустить":
            defect_value = "Пропустить"
        else:
            defect_value = self.defect_choice.get()

        data = {
            "customer_display": customer_display,
            "plate": plate_value,
            "trailer": trailer_value,
            "driver_name": self.driver_name.get().strip(),
            "defect": defect_value,
            "issued_to": self.issued_to.get().strip(),
            # ИЗМЕНЕНО: убрано поле mechanic (оставляем пустым в Excel)
            "mechanic": "",  # Механик будет ставить подпись вручную
            "vehicle_type": self.vehicle_type.get(),
            "wheel_type": self.wheel_type.get(),
            "services": self._collect_services(),
            "wheels": self.added_wheels,  # Добавляем список колес
        }
        return data

    def _build_xlsx_only(self):
        ok, msg = self._validate()
        if not ok:
            messagebox.showerror("Ошибка", msg, parent=self._form_parent)
            return
        data = self._gather_data()
        try:
            # Получаем папку дня перед созданием файла
            current_folder = get_current_day_folder()

            xlsx_path = fill_excel_only(data)
            file_count = len(list(current_folder.glob("*")))

            messagebox.showinfo(
                "Готово",
                f"Excel сформирован:\n\n{xlsx_path}\n\n"
                f"Файл сохранен в папке: {current_folder.name}\n"
                f"Всего файлов в папке: {file_count}\n\n"
                f"Открываю папку с результатами.",
                parent=self._form_parent
            )
            try:
                os.startfile(str(current_folder.resolve()))
            except Exception:
                pass
        except FileNotFoundError as e:
            messagebox.showerror("Шаблон не найден", str(e), parent=self._form_parent)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Неожиданная ошибка: {e}", parent=self._form_parent)


def main():
    app = tb.Window(themename="flatly")
    WorkOrderApp(app)
    app.mainloop()


if __name__ == "__main__":
    main()